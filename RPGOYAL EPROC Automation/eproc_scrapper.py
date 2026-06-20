import io
import os
import re
import sys
import time
import shutil
import zipfile
import datetime
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
)
from webdriver_manager.chrome import ChromeDriverManager

RUPEES_PER_CRORE = 10_000_000  # 1 crore = 1,00,00,000
PIN_CODES_REFERENCE_XLSX = "List of Pin Codes of Rajasthan.xlsx"

# ====== HELPER FUNCTION TO GET BASE PATH ======
def get_base_path():
    if getattr(sys, 'frozen', False):  # If bundled as .exe
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

base_path = get_base_path()

# ====== SETUP LOGGER WITH TIMESTAMPED FILE ======
log_dir = os.path.join(base_path, "logs", "rajasthan")
os.makedirs(log_dir, exist_ok=True)

log_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
log_file = os.path.join(log_dir, f"scraper_{log_time}.log")

logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
)
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console.setFormatter(formatter)
logging.getLogger().addHandler(console)

logging.info("Logger initialized. Starting tender scraping process.")

# ====== DRIVER INITIALIZATION ======
def init_driver(headless=True, download_dir=None):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--start-maximized")
    options.page_load_strategy = "eager"
    if download_dir:
        os.makedirs(download_dir, exist_ok=True)
        dl = os.path.abspath(download_dir)
        options.add_experimental_option(
            "prefs",
            {
                "download.default_directory": dl,
                "download.prompt_for_download": False,
                "directory_upgrade": True,
                "safebrowsing.enabled": True,
            },
        )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

# ====== NAVIGATE TO TENDER LISTING ======
def navigate_to_main_page(driver):
    driver.get("https://eproc.rajasthan.gov.in/")
    # Notice dialog is not always shown (or markup may change); do not block the run on it.
    try:
        WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "alertbutclose"))
        ).click()
        logging.info("Closed homepage notice dialog.")
    except TimeoutException:
        logging.info("No homepage notice dialog; continuing to tenders link.")
    # Link IDs on the portal are not stable across releases.
    # Prefer a text-based locator for "Tenders by Organisation".
    clicked = False
    last_err = None
    locators = [
        (
            By.XPATH,
            "//a[contains(translate(normalize-space(.),"
            " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),"
            " 'tenders by organisation')]",
        ),
        # Fallback IDs seen on some portal versions.
        (By.ID, "PageLink_11"),
        (By.ID, "PageLink_10"),
    ]
    for by, value in locators:
        try:
            WebDriverWait(driver, 12).until(
                EC.element_to_be_clickable((by, value))
            ).click()
            clicked = True
            break
        except Exception as e:
            last_err = e
            continue
    if not clicked:
        raise TimeoutException(
            f"Could not open 'Tenders by Organisation' page: {last_err}"
        )
    logging.info("Navigated to tenders-by-organisation page.")

# ====== FETCH ORGANISATIONS AND LINKS ======
def get_tenders(driver):
    return driver.find_elements(By.XPATH, "//a[contains(@id,'DirectLink')]")

# ====== FETCH INDIVIDUAL FIELD FROM TENDER PAGE ======
def get_field_text_from_popup(driver, field_name, log_miss=True):
    try:
        xpath = f"//td/b[normalize-space()='{field_name}']/ancestor::td/following-sibling::td[1]"
        element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return element.text.strip()
    except Exception as e:
        if log_miss:
            logging.warning(f"Could not extract '{field_name}': {e}")
        return "Not Found"


def get_zipcode_from_page(driver):
    """Try common NIC/eProc labels for PIN / ZIP on the tender detail table."""
    for label in (
        "Pin Code",
        "Pincode",
        "PIN Code",
        "ZIP Code",
        "Zipcode",
        "Postal Code",
    ):
        try:
            xpath = f"//td/b[normalize-space()='{label}']/ancestor::td/following-sibling::td[1]"
            element = WebDriverWait(driver, 1).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            text = element.text.strip()
            if text:
                return text
        except Exception:
            continue
    return "Not Found"


def find_download_zip_link(driver):
    """Locate 'Download a zip file' (or similar) link on the tender page."""
    xpaths = [
        "//a[contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'download a zip')]",
        "//a[contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'zip file')]",
        "//a[contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'download') and contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'zip')]",
    ]
    for xp in xpaths:
        for a in driver.find_elements(By.XPATH, xp):
            try:
                if a.is_displayed() and a.is_enabled():
                    return a
            except Exception:
                continue
    for a in driver.find_elements(By.TAG_NAME, "a"):
        try:
            txt = (a.text or "").lower()
            if "zip" in txt and "download" in txt and a.is_displayed():
                return a
        except Exception:
            continue
    return None


def find_restart_control(driver):
    """Button or link labeled Restart on NIC session/timeout pages."""
    xpaths = [
        "//a[contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'restart')]",
        "//button[contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'restart')]",
        "//input[(@type='submit' or @type='button') and contains(translate(@value, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'restart')]",
    ]
    for xp in xpaths:
        for el in driver.find_elements(By.XPATH, xp):
            try:
                if el.is_displayed() and el.is_enabled():
                    return el
            except Exception:
                continue
    return None


def recover_tender_page_from_timeout(
    driver, tender_label="", max_restarts=3, tender_url=None
):
    """
    After opening a tender URL, the portal may show a session/timeout page with Restart.
    Click Restart (up to max_restarts times). Restart often lands on the portal home page,
    so there is no zip link there — we must open tender_url again to return to the tender view.
    """
    deadline = time.time() + 60
    restarts_done = 0
    while time.time() < deadline:
        if find_download_zip_link(driver):
            return True
        restart = find_restart_control(driver)
        if restart and restarts_done < max_restarts:
            logging.info(
                "Session/timeout page — clicking Restart for %s (%s/%s)",
                tender_label or "tender",
                restarts_done + 1,
                max_restarts,
            )
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", restart
                )
                time.sleep(0.35)
                restart.click()
            except Exception as e:
                logging.warning("Restart click failed: %s", e)
            restarts_done += 1
            time.sleep(2.5)
            if tender_url:
                logging.info(
                    "Re-opening tender URL after Restart (portal often returns to home): %s",
                    tender_label or "tender",
                )
                try:
                    driver.get(tender_url)
                    time.sleep(2.5)
                except Exception as e:
                    logging.warning("Could not re-open tender URL: %s", e)
            continue
        time.sleep(0.9)
    return bool(find_download_zip_link(driver))


def sanitize_project_zip_basename(tender_data, seq_index):
    """Filesystem-safe base name from tender Title (project name), else Work Description."""
    tender_id = str(tender_data.get("Tender ID") or "").strip()
    raw = (tender_data.get("Title") or "").strip()
    if not raw or raw.lower() == "not found":
        raw = (tender_data.get("Work Description") or "").strip()
    if not raw or raw.lower() == "not found":
        raw = f"tender_{seq_index:04d}"
    if tender_id:
        raw = f"{tender_id}_{raw}"
    base = sanitize_folder_name(raw)
    if len(base) > 150:
        base = base[:150].rstrip("_")
    return base or f"tender_{seq_index:04d}"


def unique_zip_path_in_dir(target_dir, basename):
    """Return full path .../basename.zip, or basename_2.zip, etc. if needed."""
    p = os.path.join(target_dir, f"{basename}.zip")
    if not os.path.exists(p):
        return p
    n = 2
    while True:
        p = os.path.join(target_dir, f"{basename}_{n}.zip")
        if not os.path.exists(p):
            return p
        n += 1


def _safe_click_zip_link(driver, link):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
        time.sleep(0.3)
        link.click()
        return True
    except Exception as e:
        logging.warning("Zip link click failed: %s", e)
        return False


def wait_for_zip_in_folder(download_dir, before_names, timeout=180):
    """Wait until a new .zip appears and .crdownload files are gone."""
    deadline = time.time() + timeout
    before_names = set(before_names)
    while time.time() < deadline:
        try:
            names = os.listdir(download_dir)
        except OSError:
            time.sleep(0.4)
            continue
        if any(n.endswith(".crdownload") for n in names):
            time.sleep(0.5)
            continue
        zips = [
            n
            for n in names
            if n.endswith(".zip") and n not in before_names
        ]
        if zips:
            zips.sort(key=lambda n: os.path.getmtime(os.path.join(download_dir, n)), reverse=True)
            return os.path.join(download_dir, zips[0])
        time.sleep(0.4)
    return None


_captcha_callback = None

# BOQ user actions (CAPTCHA / stuck page / failed download)
BOQ_ACTION_CONTINUE = "continue"
BOQ_ACTION_SKIP = "skip"
BOQ_ACTION_RESTART = "restart"
BOQ_RESTART_SENTINEL = "__BOQ_RESTART_BROWSER__"


def _normalize_boq_user_action(value):
    if value is True or value == BOQ_ACTION_CONTINUE:
        return BOQ_ACTION_CONTINUE
    if value == BOQ_ACTION_RESTART:
        return BOQ_ACTION_RESTART
    return BOQ_ACTION_SKIP


def _prompt_boq_user_action_cli(label, reason=""):
    reason_line = f"  Reason: {reason}\n" if reason else ""
    print(
        f"\n  >>> BOQ help needed — {label}\n"
        f"{reason_line}"
        "  [Enter] Continue — fix page/CAPTCHA, retry in same browser\n"
        "  [*]     Restart — close Chrome, fresh window, retry THIS tender\n"
        "  [s]     Skip this tender\n"
    )
    try:
        raw = input("  Choice [Enter/* / s]: ").strip().lower()
    except EOFError:
        return BOQ_ACTION_SKIP
    if raw in ("*", "r", "restart"):
        return BOQ_ACTION_RESTART
    if raw in ("s", "skip"):
        return BOQ_ACTION_SKIP
    return BOQ_ACTION_CONTINUE


def _ask_boq_user_action(label, reason=""):
    global _captcha_callback
    if _captcha_callback is not None:
        msg = f"{reason} — {label}" if reason else label
        return _normalize_boq_user_action(_captcha_callback(msg))
    return _prompt_boq_user_action_cli(label, reason)


def _restart_boq_driver(driver, staging_dir):
    try:
        driver.quit()
    except Exception:
        pass
    time.sleep(1)
    logging.info("Restarting Chrome for BOQ downloads (new browser session).")
    return init_driver(headless=False, download_dir=staging_dir)


def try_download_tender_boq_zip(driver, download_dir, tender_label, tender_url=None):
    """
    Click “Download a zip file”. If a .zip appears quickly, no CAPTCHA step.
    Otherwise prompt for CAPTCHA / page ready, click again (up to 3 extra tries).
    tender_url is used after Restart recovery (home page → open tender again).
    Returns path to downloaded .zip inside download_dir, or None.
    """
    link = find_download_zip_link(driver)
    if not link:
        logging.warning("No zip download link for: %s", tender_label)
        return None

    before_first = set(os.listdir(download_dir))
    if not _safe_click_zip_link(driver, link):
        return None

    path = wait_for_zip_in_folder(download_dir, before_first, timeout=25)
    if path:
        logging.info("Zip downloaded without CAPTCHA for: %s", tender_label)
        return path

    for extra in range(3):
        action = _ask_boq_user_action(
            tender_label,
            reason=f"CAPTCHA or download retry ({extra + 1}/3)",
        )
        if action == BOQ_ACTION_RESTART:
            return BOQ_RESTART_SENTINEL
        if action == BOQ_ACTION_SKIP:
            logging.info("User skipped BOQ for: %s", tender_label)
            return None

        recover_tender_page_from_timeout(
            driver, tender_label, max_restarts=3, tender_url=tender_url
        )
        link2 = find_download_zip_link(driver)
        if not link2:
            logging.warning("Zip link not found after wait (round %s): %s", extra + 1, tender_label)
            continue

        before_round = set(os.listdir(download_dir))
        if not _safe_click_zip_link(driver, link2):
            continue

        path = wait_for_zip_in_folder(download_dir, before_round, timeout=180)
        if path:
            return path
        logging.warning("No .zip after click (round %s): %s", extra + 1, tender_label)

    logging.warning("Giving up on zip download for: %s", tender_label)
    return None


def extract_boq_from_zip_to_csv(zip_path, out_csv_path):
    """
    Find BOQ spreadsheet inside the tender zip (name contains 'boq' or first .xlsx/.xls/.csv).
    Write a flattened CSV (all sheets stacked with _sheet column for Excel).
    """
    with zipfile.ZipFile(zip_path, "r") as zf:
        members = [n for n in zf.namelist() if not n.endswith("/")]
        if not members:
            raise ValueError("Empty zip")
        base_names = [(n, os.path.basename(n)) for n in members]
        boq = [n for n, b in base_names if "boq" in b.lower()]
        candidates = boq or [
            n
            for n, b in base_names
            if b.lower().endswith((".xlsx", ".xls", ".csv"))
        ]
        if not candidates:
            raise ValueError("No spreadsheet/BOQ file in zip")
        inner = candidates[0]
        ext = os.path.splitext(inner)[1].lower()
        raw = zf.read(inner)

    bio = io.BytesIO(raw)
    if ext == ".csv":
        df = pd.read_csv(bio, header=None, encoding_errors="replace")
        df.insert(0, "_source_file", os.path.basename(inner))
        df.to_csv(out_csv_path, index=False)
        return

    if ext == ".xlsx":
        xl = pd.ExcelFile(io.BytesIO(raw), engine="openpyxl")
    elif ext == ".xls":
        try:
            xl = pd.ExcelFile(io.BytesIO(raw), engine="xlrd")
        except ImportError as e:
            raise ImportError(
                "Reading .xls BOQ requires xlrd. Run: pip install xlrd"
            ) from e
    else:
        xl = pd.ExcelFile(io.BytesIO(raw))

    parts = []
    for sheet in xl.sheet_names:
        sh = pd.read_excel(xl, sheet_name=sheet, header=None)
        sh.insert(0, "_sheet", sheet)
        sh.insert(0, "_source_file", os.path.basename(inner))
        parts.append(sh)
    combined = pd.concat(parts, ignore_index=True)
    combined.to_csv(out_csv_path, index=False)


# ----- BOQ raw CSV -> cleaned columns + summary CSV + PDF -----

_BOQ_COLUMN_TARGETS = [
    (
        "Sl. No.",
        [
            "sl. no",
            "sl no",
            "s.no",
            "s. no",
            "serial no",
            "sr no",
            "sr. no",
            "srno",
            "slno",
        ],
    ),
    (
        "Item Description",
        [
            "item description",
            "description of item",
            "description",
            "item desc",
            "name of item",
            "particulars",
        ],
    ),
    (
        "Quantity",
        ["quantity", "qty", "qty."],
    ),
    (
        "Units",
        ["units", "unit", "uom", "u.o.m"],
    ),
    (
        "Estimated Rate",
        [
            "estimated rate",
            "unit rate",
            "basic rate",
            "rate",
        ],
    ),
    (
        "TOTAL AMOUNT With Taxes",
        [
            "total amount with taxes",
            "total amt with taxes",
            "amount with taxes",
            "total with tax",
            "total amount (with taxes)",
        ],
    ),
    (
        "TOTAL AMOUNT In Words",
        [
            "total amount in words",
            "amount in words",
            "in words",
            "rupees in words",
            "total in words",
        ],
    ),
]


def _norm_hdr_cell(val):
    s = str(val).strip() if pd.notna(val) else ""
    return re.sub(r"\s+", " ", s.lower())


def _row_looks_like_boq_dtype_row(row_series):
    """Row where cells are like Number #, Text # (NIC BOQ layout)."""
    pat = re.compile(r"^(number|text|date)\s*#", re.I)
    hits = 0
    for x in row_series:
        if pd.isna(x):
            continue
        t = str(x).strip()
        if pat.match(t):
            hits += 1
    return hits >= 2


def _pick_column_index(header_cells, keywords):
    """Best-matching column index for header row; keywords longest-first."""
    headers = [str(c).strip() if pd.notna(c) else "" for c in header_cells]
    best_j = None
    best_len = 0
    for j, h in enumerate(headers):
        hn = _norm_hdr_cell(h)
        if not hn:
            continue
        for kw in sorted(keywords, key=len, reverse=True):
            kn = _norm_hdr_cell(kw)
            if not kn:
                continue
            if kn in hn or hn in kn:
                if len(kn) > best_len:
                    best_len = len(kn)
                    best_j = j
    return best_j


def _parse_numeric_for_summary(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    digits = re.sub(r"[^\d.\-]", "", s)
    if not digits or digits in (".", "-"):
        return None
    try:
        return float(digits)
    except ValueError:
        return None


def _is_valid_sl_no_value(val):
    """
    Keep rows only when Sl. No. is:
      - numeric / decimal (e.g., 1, 1.1, 2.01), or
      - 'Total in Figures' (case-insensitive).
    """
    if _is_blank_boq_cell(val):
        return False
    s = str(val).strip()
    s_norm = re.sub(r"\s+", " ", s).lower()
    if s_norm == "total in figures":
        return True
    return re.fullmatch(r"\d+(?:\.\d+)?", s) is not None


def _is_integer_only_text(val):
    if _is_blank_boq_cell(val):
        return False
    s = str(val).strip()
    return re.fullmatch(r"\d+", s) is not None


def _row_all_non_empty_values_are_integers(row):
    vals = []
    for v in row:
        if _is_blank_boq_cell(v):
            continue
        vals.append(str(v).strip())
    if not vals:
        return False
    return all(_is_integer_only_text(v) for v in vals)


def _extract_boq_blocks_from_raw_df(raw_df):
    """
    Yield (header_names list aligned to ncol, data_subframe with same ncol) for each BOQ table.
    """
    ncols = raw_df.shape[1]
    start = 0
    while start < len(raw_df):
        dtype_idx = None
        for i in range(start, len(raw_df)):
            if _row_looks_like_boq_dtype_row(raw_df.iloc[i]):
                dtype_idx = i
                break
        if dtype_idx is None:
            break
        header_idx = dtype_idx + 1
        if header_idx >= len(raw_df):
            break
        header_row = raw_df.iloc[header_idx]
        names = [
            str(header_row.iloc[j]).strip() if pd.notna(header_row.iloc[j]) else f"col_{j}"
            for j in range(ncols)
        ]
        data_start = header_idx + 1
        data_end = len(raw_df)
        for i in range(data_start, len(raw_df)):
            if _row_looks_like_boq_dtype_row(raw_df.iloc[i]):
                data_end = i
                break
        block = raw_df.iloc[data_start:data_end].copy()
        block.columns = [f"c{j}" for j in range(ncols)]
        yield names, block
        start = data_end


def _build_clean_boq_dataframe(header_names, block_df):
    """Map target columns to extracted series; align by original column index."""
    ncol = len(header_names)
    if block_df.shape[1] != ncol:
        return None
    colmap = {}
    for out_name, keywords in _BOQ_COLUMN_TARGETS:
        j = _pick_column_index(header_names, keywords)
        if j is not None:
            colmap[out_name] = block_df.iloc[:, j].astype(str).str.strip()
        else:
            colmap[out_name] = pd.Series([""] * len(block_df))
    out = pd.DataFrame(colmap)
    mask = out.apply(lambda r: any(str(x).strip() for x in r), axis=1)
    out = out.loc[mask].reset_index(drop=True)
    # Drop rows where every non-empty value is integer-only (e.g. helper/index rows).
    if not out.empty:
        out = out[
            ~out.apply(_row_all_non_empty_values_are_integers, axis=1)
        ].reset_index(drop=True)
    if "Sl. No." in out.columns:
        out = out[out["Sl. No."].apply(_is_valid_sl_no_value)].reset_index(drop=True)
    # Drop the common "column number" row found in many BOQ sheets,
    # e.g. first data row values like 1,2,3,4,... across columns.
    if not out.empty:
        first = out.iloc[0]
        vals = [str(first.get(c, "")).strip() for c in out.columns]
        nums = []
        for v in vals:
            if re.fullmatch(r"\d+", v):
                nums.append(int(v))
            elif _is_blank_boq_cell(v):
                continue
            else:
                nums = []
                break
        if len(nums) >= 4 and nums == list(range(1, len(nums) + 1)):
            out = out.iloc[1:].reset_index(drop=True)
    return out


def _is_section_header_row_for_pdf(row_dict):
    """
    A section header row is:
      - Sl. No. is integer (not decimal), and
      - Quantity, Units, Estimated Rate, TOTAL AMOUNT With Taxes are all blank.
    """
    sl = str(row_dict.get("Sl. No.", "")).strip()
    if not re.fullmatch(r"\d+", sl):
        return False
    required_blank = [
        "Quantity",
        "Units",
        "Estimated Rate",
        "TOTAL AMOUNT With Taxes",
    ]
    for c in required_blank:
        if not _is_blank_boq_cell(row_dict.get(c, "")):
            return False
    return True


def _summarise_boq_clean_df(df):
    """Return list of (label, value) for summary section."""
    lines = []
    n = len(df)
    lines.append(("Line items (rows)", str(n)))
    qcol = "Quantity"
    if qcol in df.columns:
        qs = [_parse_numeric_for_summary(x) for x in df[qcol]]
        qs = [x for x in qs if x is not None]
        if qs:
            lines.append(("Sum of Quantity", f"{sum(qs):,.2f}".rstrip("0").rstrip(".")))
    rcol = "Estimated Rate"
    tcol = "TOTAL AMOUNT With Taxes"
    if tcol in df.columns:
        ts = [_parse_numeric_for_summary(x) for x in df[tcol]]
        ts = [x for x in ts if x is not None]
        if ts:
            lines.append(
                ("Sum of TOTAL AMOUNT With Taxes", f"{sum(ts):,.2f}".rstrip("0").rstrip("."))
            )
    if rcol in df.columns:
        rs = [_parse_numeric_for_summary(x) for x in df[rcol]]
        rs = [x for x in rs if x is not None]
        if rs:
            lines.append(("Min Estimated Rate", f"{min(rs):,.2f}"))
            lines.append(("Max Estimated Rate", f"{max(rs):,.2f}"))
    return lines


def _is_blank_boq_cell(val):
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    s = str(val).strip().lower()
    return s in ("", "nan", "none", "nat", "<na>")


def _prepare_boq_dataframe_for_output(df):
    """Replace NaN / 'nan' with empty strings for CSV + PDF (totals-only columns stay blank)."""
    out = df.copy()
    for c in out.columns:
        out[c] = out[c].apply(
            lambda x: "" if _is_blank_boq_cell(x) else str(x).strip()
        )
    return out


def _assign_tender_ids(tenders):
    """Assign stable per-organisation IDs used in Excel and BOQ file names."""
    for i, t in enumerate(tenders, start=1):
        if not str(t.get("Tender ID") or "").strip():
            t["Tender ID"] = f"T{i:04d}"


def _safe_pdf_text(s):
    if s is None:
        return ""
    t = str(s)
    return t.encode("latin-1", "replace").decode("latin-1")


def _pdf_paragraph_markup(val):
    """XML-safe text for ReportLab Paragraph; blank cells become minimal space."""
    import html as html_module

    if _is_blank_boq_cell(val):
        return " "
    s = str(val).strip()
    s = html_module.escape(s)
    return s.replace("\n", "<br/>")


def _boq_pdf_column_widths(headers, usable_width_pt):
    """Wider columns for description / amounts / words; narrow for Sl / Qty / Units."""
    weights = []
    for h in headers:
        hn = str(h).lower()
        if "description" in hn:
            weights.append(3.4)
        elif "words" in hn:
            weights.append(2.2)
        elif "taxes" in hn or ("total" in hn and "amount" in hn):
            weights.append(1.5)
        elif "rate" in hn:
            weights.append(1.1)
        elif "quantity" in hn or "qty" in hn:
            weights.append(0.75)
        elif "unit" in hn:
            weights.append(0.65)
        elif "sl" in hn or "no" in hn:
            weights.append(0.55)
        else:
            weights.append(1.0)
    s = sum(weights)
    return [usable_width_pt * (w / s) for w in weights]


def _write_boq_pdf_report(section_tables, summary_pairs, pdf_path, title):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
    except ImportError as e:
        raise ImportError("Install reportlab for BOQ PDF: pip install reportlab") from e

    LM = RM = 14 * mm
    TM = BM = 12 * mm
    page_w_pt, page_h_pt = landscape(A4)
    usable_w = page_w_pt - LM - RM

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        leftMargin=LM,
        rightMargin=RM,
        topMargin=TM,
        bottomMargin=BM,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name="BoqTitle",
        parent=styles["Title"],
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=10,
    )
    cell_style = ParagraphStyle(
        name="BoqCell",
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        leftIndent=3,
        rightIndent=3,
        spaceBefore=2,
        spaceAfter=2,
        alignment=TA_LEFT,
    )
    header_cell_style = ParagraphStyle(
        name="BoqHeadCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.whitesmoke,
    )

    flow = []
    flow.append(
        Paragraph(_safe_pdf_text(title or "BOQ Report"), title_style)
    )
    flow.append(Spacer(1, 4))

    if not section_tables:
        flow.append(Paragraph("No data rows extracted.", styles["Normal"]))
        doc.build(flow)
        return

    for sec_idx, (sec_name, clean_df) in enumerate(section_tables, start=1):
        if clean_df is None or clean_df.empty:
            continue
        if sec_idx > 1:
            flow.append(PageBreak())
        flow.append(
            Paragraph(
                _safe_pdf_text(f"Table {sec_idx}: {sec_name}"),
                ParagraphStyle(
                    name=f"BoqSec_{sec_idx}",
                    parent=styles["Heading3"],
                    fontSize=11,
                    leading=13,
                    textColor=colors.HexColor("#1F4E79"),
                    spaceAfter=6,
                ),
            )
        )
        headers = list(clean_df.columns)
        col_widths = _boq_pdf_column_widths(headers, usable_w)
        header_row = [
            Paragraph(_pdf_paragraph_markup(h), header_cell_style) for h in headers
        ]
        table_rows = [header_row]
        span_rows = []
        full_span_rows = []
        sl_col_idx = headers.index("Sl. No.") if "Sl. No." in headers else None
        est_col_idx = headers.index("Estimated Rate") if "Estimated Rate" in headers else None
        item_desc_idx = headers.index("Item Description") if "Item Description" in headers else None
        section_header_style = ParagraphStyle(
            name=f"BoqSectionRow_{sec_idx}",
            parent=cell_style,
            fontName="Helvetica-Bold",
            alignment=1,  # center
        )
        for _, row in clean_df.iterrows():
            row_dict = {c: row[c] for c in headers}
            if _is_section_header_row_for_pdf(row_dict) and item_desc_idx is not None:
                section_txt = _pdf_paragraph_markup(row_dict.get("Item Description", ""))
                row_cells = [Paragraph(" ", cell_style) for _ in headers]
                row_cells[0] = Paragraph(section_txt or " ", section_header_style)
                table_rows.append(row_cells)
                full_span_rows.append(len(table_rows) - 1)
            else:
                row_cells = [Paragraph(_pdf_paragraph_markup(row[c]), cell_style) for c in headers]
                table_rows.append(row_cells)
                if sl_col_idx is not None and est_col_idx is not None and est_col_idx >= sl_col_idx:
                    sl_val_norm = re.sub(r"\s+", " ", str(row["Sl. No."]).strip()).lower()
                    if sl_val_norm == "total in figures":
                        # Merge label from 'Sl. No.' through 'Estimated Rate' in PDF row.
                        span_rows.append(len(table_rows) - 1)
        t = Table(
            table_rows,
            colWidths=col_widths,
            repeatRows=1,
            splitByRow=1,
            spaceBefore=6,
            spaceAfter=10,
        )
        table_style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B0B0B0")),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [colors.white, colors.HexColor("#F0F5FA")],
            ),
        ]
        if sl_col_idx is not None and est_col_idx is not None and est_col_idx >= sl_col_idx:
            for r_idx in span_rows:
                table_style_cmds.extend(
                    [
                        ("SPAN", (sl_col_idx, r_idx), (est_col_idx, r_idx)),
                        ("ALIGN", (sl_col_idx, r_idx), (est_col_idx, r_idx), "LEFT"),
                        ("VALIGN", (sl_col_idx, r_idx), (est_col_idx, r_idx), "MIDDLE"),
                        ("FONTNAME", (sl_col_idx, r_idx), (-1, r_idx), "Helvetica-Bold"),
                        ("BACKGROUND", (sl_col_idx, r_idx), (-1, r_idx), colors.HexColor("#FFF2CC")),
                    ]
                )
        for r_idx in full_span_rows:
            table_style_cmds.extend(
                [
                    ("SPAN", (0, r_idx), (-1, r_idx)),
                    ("ALIGN", (0, r_idx), (-1, r_idx), "CENTER"),
                    ("VALIGN", (0, r_idx), (-1, r_idx), "MIDDLE"),
                ]
            )
        t.setStyle(TableStyle(table_style_cmds))
        flow.append(t)
    doc.build(flow)


def process_and_export_clean_boq(raw_csv_path, project_title="", pdf_output_dir=None):
    """
    Read raw BOQ CSV (no header), find dtype row -> header -> data, extract requested
    columns, append summary rows, write <base>_cleaned.csv next to raw file.
    PDF goes to pdf_output_dir if set (e.g. BOQs/All BOQs <Dept>/), else beside the CSV.
    """
    if not os.path.isfile(raw_csv_path):
        return None, None
    raw = pd.read_csv(raw_csv_path, header=None, encoding_errors="replace")
    if raw.empty:
        logging.warning("Empty BOQ raw file: %s", raw_csv_path)
        return None, None

    parts = []
    section_tables = []
    for header_names, block in _extract_boq_blocks_from_raw_df(raw):
        clean = _build_clean_boq_dataframe(header_names, block)
        if clean is not None and not clean.empty:
            parts.append(clean)
            section_tables.append((f"BOQ Section {len(section_tables) + 1}", clean))

    if not parts:
        logging.warning("No BOQ table (dtype/header) found in: %s", raw_csv_path)
        return None, None

    merged = pd.concat(parts, ignore_index=True)

    base = os.path.splitext(raw_csv_path)[0]
    cleaned_csv = f"{base}_cleaned.csv"
    pdf_name = os.path.basename(base) + "_report.pdf"
    if pdf_output_dir:
        os.makedirs(pdf_output_dir, exist_ok=True)
        pdf_path = os.path.join(pdf_output_dir, pdf_name)
    else:
        pdf_path = f"{base}_report.pdf"

    summary_pairs = _summarise_boq_clean_df(merged)
    merged_out = _prepare_boq_dataframe_for_output(merged)

    merged_out.to_csv(cleaned_csv, index=False)
    sum_df = pd.DataFrame(summary_pairs, columns=["Metric", "Value"])
    with open(cleaned_csv, "a", encoding="utf-8", newline="") as f:
        f.write("\n")
    sum_df.to_csv(cleaned_csv, mode="a", index=False)

    title = (project_title or os.path.basename(base)).strip()
    try:
        _write_boq_pdf_report(section_tables, summary_pairs, pdf_path, title)
        logging.info("BOQ PDF report: %s", pdf_path)
    except Exception as e:
        logging.warning("BOQ PDF export failed: %s", e)

    logging.info("BOQ cleaned CSV: %s", cleaned_csv)
    return cleaned_csv, pdf_path


def write_boq_summary_csv(tenders, path):
    rows = []
    for t in tenders:
        rows.append(
            {
                "Title": t.get("Title", ""),
                "Tender Link": t.get("Tender Link", ""),
                "BOQ_Status": t.get("BOQ_Status", ""),
                "BOQ_ZIP": t.get("BOQ_ZIP", ""),
                "BOQ_CSV": t.get("BOQ_CSV", ""),
                "BOQ_Cleaned_CSV": t.get("BOQ_Cleaned_CSV", ""),
                "BOQ_Report_PDF": t.get("BOQ_Report_PDF", ""),
            }
        )
    pd.DataFrame(rows).to_csv(path, index=False)


def download_boq_for_tenders(tenders, boq_dept_dir, driver, staging_dir, department_name=""):
    """
    For each tender: open URL, download zip via staging_dir (Chrome downloads there),
    move/rename zip to boq_dept_dir using project (Title) name, extract BOQ CSV.
    PDF reports are written under BOQs/All BOQs <department>/ (same parent as boq_dept_dir).
    Returns the active WebDriver (may be a new instance after user-requested restart).
    Mutates tender BOQ_* fields.
    """
    os.makedirs(boq_dept_dir, exist_ok=True)
    boqs_parent = os.path.dirname(boq_dept_dir)
    dept_label = (department_name or os.path.basename(boq_dept_dir)).strip()
    all_boqs_pdf_dir = os.path.join(
        boqs_parent, sanitize_folder_name(f"All BOQs {dept_label}")
    )
    os.makedirs(all_boqs_pdf_dir, exist_ok=True)
    for t in tenders:
        t.setdefault("BOQ_Status", "")
        t.setdefault("BOQ_ZIP", "")
        t.setdefault("BOQ_CSV", "")
        t.setdefault("BOQ_Cleaned_CSV", "")
        t.setdefault("BOQ_Report_PDF", "")

    total = len(tenders)
    for i, tender_data in enumerate(tenders, start=1):
        link = tender_data.get("Tender Link")
        if not link:
            tender_data["BOQ_Status"] = "no_link"
            continue

        label = f"{i}/{total} — {str(tender_data.get('Title', ''))[:80]}"
        tender_done = False
        while not tender_done:
            try:
                driver.get(link)
                time.sleep(1.5)
            except Exception as e:
                logging.error("Open tender failed %s: %s", i, e)
                action = _ask_boq_user_action(label, "Could not open tender page")
                if action == BOQ_ACTION_RESTART:
                    driver = _restart_boq_driver(driver, staging_dir)
                    continue
                if action == BOQ_ACTION_SKIP:
                    tender_data["BOQ_Status"] = "page_error"
                    tender_done = True
                    continue
                continue

            if not recover_tender_page_from_timeout(
                driver, label, max_restarts=3, tender_url=link
            ):
                logging.warning(
                    "Tender page did not load (no zip link after Restart) for: %s",
                    label,
                )
                action = _ask_boq_user_action(
                    label, "Page stuck / no download link (timeout)"
                )
                if action == BOQ_ACTION_RESTART:
                    driver = _restart_boq_driver(driver, staging_dir)
                    continue
                if action == BOQ_ACTION_SKIP:
                    tender_data["BOQ_Status"] = "timeout_or_no_zip_link"
                    tender_done = True
                    continue
                continue

            zip_path = try_download_tender_boq_zip(
                driver, staging_dir, label, tender_url=link
            )
            if zip_path == BOQ_RESTART_SENTINEL:
                driver = _restart_boq_driver(driver, staging_dir)
                continue
            if not zip_path:
                action = _ask_boq_user_action(label, "ZIP download failed")
                if action == BOQ_ACTION_RESTART:
                    driver = _restart_boq_driver(driver, staging_dir)
                    continue
                if action == BOQ_ACTION_SKIP:
                    tender_data["BOQ_Status"] = "download_failed"
                    tender_done = True
                    continue
                continue

            proj_base = sanitize_project_zip_basename(tender_data, i)
            dest_zip = unique_zip_path_in_dir(boq_dept_dir, proj_base)
            try:
                if os.path.abspath(zip_path) != os.path.abspath(dest_zip):
                    if os.path.isfile(dest_zip):
                        os.remove(dest_zip)
                    shutil.move(zip_path, dest_zip)
                zip_final = dest_zip
            except OSError as e:
                logging.warning(
                    "Could not move zip to %s (%s); using staging path", dest_zip, e
                )
                zip_final = zip_path

            csv_out = f"{os.path.splitext(zip_final)[0]}_boq.csv"
            try:
                extract_boq_from_zip_to_csv(zip_final, csv_out)
                tender_data["BOQ_Status"] = "ok"
                tender_data["BOQ_ZIP"] = zip_final
                tender_data["BOQ_CSV"] = csv_out
                logging.info("BOQ extracted: %s", csv_out)
                try:
                    ccsv, pdfp = process_and_export_clean_boq(
                        csv_out,
                        tender_data.get("Title", ""),
                        pdf_output_dir=all_boqs_pdf_dir,
                    )
                    if ccsv:
                        tender_data["BOQ_Cleaned_CSV"] = ccsv
                    if pdfp:
                        tender_data["BOQ_Report_PDF"] = pdfp
                except Exception as clean_ex:
                    logging.warning("BOQ clean/PDF step failed: %s", clean_ex)
            except Exception as ex:
                tender_data["BOQ_Status"] = f"extract_failed: {ex}"
                tender_data["BOQ_ZIP"] = zip_final
                logging.error("BOQ extract failed tender %s: %s", i, ex)
            tender_done = True

    summary_path = os.path.join(boq_dept_dir, "boq_download_summary.csv")
    try:
        write_boq_summary_csv(tenders, summary_path)
        logging.info("BOQ summary: %s", summary_path)
    except Exception as e:
        logging.warning("Could not write BOQ summary: %s", e)
    return driver


# ====== EXTRACT FIELDS FOR ONE TENDER ======
def get_org_tenders(org_name, org_id):
    driver = init_driver(headless=True)
    navigate_to_main_page(driver)

    try:
        org_element = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, org_id))
        )
        # Some portal pages overlay a right-aligned table cell over links.
        # Use a click fallback sequence to avoid intermittent intercept errors.
        clicked = False
        click_err = None
        for _ in range(3):
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center', inline:'nearest'});",
                    org_element,
                )
                time.sleep(0.25)
                WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.ID, org_id))
                ).click()
                clicked = True
                break
            except ElementClickInterceptedException as e:
                click_err = e
                try:
                    driver.execute_script("arguments[0].click();", org_element)
                    clicked = True
                    break
                except Exception as js_e:
                    click_err = js_e
                    time.sleep(0.35)
            except Exception as e:
                click_err = e
                time.sleep(0.35)
        if not clicked:
            raise click_err or RuntimeError("Organisation click failed")
        logging.info(f"Selected organisation: {org_name}")
    except Exception as e:
        logging.error(f"Failed to click organisation {org_name}: {e}")
        driver.quit()
        return []

    tender_links = [t.get_attribute("href") for t in get_tenders(driver)]
    total = len(tender_links)
    logging.info(f"Found {total} tenders for organisation: {org_name}")

    fields_to_extract = [
        "Title", "Work Description", "Tender Value in ₹",
        "Location", "Organisation Chain", "Bid Submission End Date"
    ]

    tenders = []
    for i, link in enumerate(tender_links, start=1):
        try:
            driver.get(link)
            tender_data = {
                "Organisation": org_name,
                "Tender Link": link,
            }
            for field in fields_to_extract:
                tender_data[field] = get_field_text_from_popup(driver, field)
            tender_data["Zipcode"] = get_zipcode_from_page(driver)

            tenders.append(tender_data)
            remaining = total - i
            percent = round((i / total) * 100, 2)
            logging.info(f"[{i}/{total}] - Recorded tender: {tender_data['Title']} | Remaining: {remaining} | {percent}% done")

        except Exception as e:
            logging.error(f"Error processing tender {i}/{total}: {e}")
            continue

    driver.quit()
    logging.info(f"Completed scraping for organisation: {org_name}")
    return tenders

def extract_orgs_and_anchor_ids_selenium(driver):
    org_dict = {}

    # Portal markup occasionally changes: table id may be absent.
    # Prefer waiting for DirectLink anchors and then resolve row/cell context.
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//a[contains(@id,'DirectLink')]"))
        )
    except TimeoutException:
        logging.warning("Organisation page loaded but no DirectLink anchors found.")
        return org_dict

    anchors = driver.find_elements(By.XPATH, "//a[contains(@id,'DirectLink')]")
    for anchor in anchors:
        try:
            anchor_id = (anchor.get_attribute("id") or "").strip()
            if not anchor_id.startswith("DirectLink"):
                continue

            row = anchor.find_element(By.XPATH, "./ancestor::tr[1]")
            classes = (row.get_attribute("class") or "").lower()
            if "list_header" in classes:
                continue

            cells = row.find_elements(By.TAG_NAME, "td")
            org_name = ""
            if len(cells) >= 2:
                # Typical layout: [index, organisation name, tender count link]
                org_name = cells[1].text.strip()
            if not org_name:
                # Fallback: row text excluding anchor text chunk
                row_text = row.text.strip()
                anchor_text = (anchor.text or "").strip()
                if row_text:
                    org_name = row_text.replace(anchor_text, "").strip()

            if org_name:
                org_dict[org_name] = anchor_id
        except NoSuchElementException:
            continue
        except Exception:
            continue

    return org_dict


def sanitize_folder_name(name):
    """Safe folder / file fragment from organisation name."""
    name = (name or "").strip()
    for c in '<>:"/\\|?*\n\r\t':
        name = name.replace(c, "_")
    name = re.sub(r"_+", "_", name).strip("._ ")
    return name or "department"


def resolve_org_choice(org_dict, user_input):
    """
    Match user input to one organisation.
    Returns ("ok", (name, anchor_id)) | ("ambiguous", list) | ("none", None) | ("empty", None)
    """
    user_input = (user_input or "").strip()
    if not user_input:
        return "empty", None
    if user_input in org_dict:
        return "ok", (user_input, org_dict[user_input])
    lower_map = {k.lower(): k for k in org_dict}
    ul = user_input.lower()
    if ul in lower_map:
        k = lower_map[ul]
        return "ok", (k, org_dict[k])
    matches = [(k, org_dict[k]) for k in org_dict if ul in k.lower()]
    if len(matches) == 1:
        return "ok", matches[0]
    if len(matches) > 1:
        return "ambiguous", matches
    return "none", None


def resolve_org_list(org_dict, raw_line):
    """
    Parse comma-separated organisation names; each token must match exactly one org.
    Returns ("ok", [(canonical_name, anchor_id), ...]) with duplicates removed,
    or ("error", message string).
    """
    raw_line = (raw_line or "").strip()
    if not raw_line:
        return "error", "Empty input."
    tokens = [t.strip() for t in raw_line.split(",") if t.strip()]
    if not tokens:
        return "error", "No organisation names found (use commas between names)."
    resolved = []
    seen = set()
    for token in tokens:
        status, payload = resolve_org_choice(org_dict, token)
        if status == "ambiguous":
            choices = "; ".join(n for n, _ in payload)
            return "error", (
                f"'{token}' matches multiple organisations — be more specific:\n  {choices}"
            )
        if status == "none":
            return "error", f"No organisation matched '{token}'. Type 'list' to see names."
        if status == "empty":
            continue
        org_name, org_id = payload
        if org_name not in seen:
            seen.add(org_name)
            resolved.append((org_name, org_id))
    if not resolved:
        return "error", "No valid organisations to scrape."
    return "ok", resolved


def parse_inr_amount_to_rupees(text):
    """
    Parse tender value strings like '21,50,00,000' into integer rupees.
    Returns None if missing or not parseable.
    """
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return None
    s = str(text).strip()
    if not s or s.lower() == "not found":
        return None
    digits = re.sub(r"[^\d]", "", s)
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def filter_tenders_by_crore_range(tenders, min_rupees, max_rupees):
    """
    Keep tenders whose 'Tender Value in ₹' is within [min_rupees, max_rupees] inclusive.
    Excludes missing or unparseable amounts when filtering.
    """
    kept = []
    dropped_unparsed = 0
    dropped_range = 0
    for t in tenders:
        raw = t.get("Tender Value in ₹", "")
        val = parse_inr_amount_to_rupees(raw)
        if val is None:
            dropped_unparsed += 1
            continue
        if val < min_rupees or val > max_rupees:
            dropped_range += 1
            continue
        kept.append(t)
    if dropped_unparsed:
        logging.info(
            "Value filter: excluded %s tender(s) (missing or unparseable amount)",
            dropped_unparsed,
        )
    if dropped_range:
        logging.info(
            "Value filter: excluded %s tender(s) (outside crore range)",
            dropped_range,
        )
    return kept


def prompt_tender_value_filter():
    """
    Ask for 'all' (no filter) or 'lower-upper' in crore, e.g. 8-50 → ₹8 Cr to ₹50 Cr inclusive.
    Returns ('all', None, None) or ('range', min_rupees, max_rupees).
    """
    print(
        "\nTender value filter:\n"
        "  • Type  all  — keep every scraped tender (no value filter)\n"
        "  • Or a range in crore:  8-50  means ₹8 crore to ₹50 crore (inclusive)\n"
    )
    while True:
        raw = input("Tender value (all or e.g. 8-50): ").strip()
        s = raw.lower().replace(" ", "")
        if s in ("all", "*", "any", "everything"):
            return "all", None, None
        m = re.match(r"^(\d+(?:\.\d+)?)-(\d+(?:\.\d+)?)$", s)
        if not m:
            print("Invalid input. Use  all  or two numbers in crore separated by a hyphen, e.g.  8-50\n")
            continue
        lo, hi = float(m.group(1)), float(m.group(2))
        if lo <= 0 or hi <= 0:
            print("Crore limits must be positive.\n")
            continue
        if lo > hi:
            lo, hi = hi, lo
        min_r = int(round(lo * RUPEES_PER_CRORE))
        max_r = int(round(hi * RUPEES_PER_CRORE))
        if min_r > max_r:
            min_r, max_r = max_r, min_r
        return "range", min_r, max_r


def prompt_download_boq():
    print(
        "\nDownload tender ZIP files and extract BOQ to CSV?\n"
        "One Chrome window is used for every department you selected. If the zip downloads immediately,\n"
        "no CAPTCHA step runs; otherwise you’ll be asked to solve CAPTCHA and press Enter (again if it reappears).\n"
        "ZIPs are named from each tender’s title (project name) under BOQs/<department>/.\n"
    )
    s = input("Download BOQs? [y/N]: ").strip().lower()
    return s in ("y", "yes", "1", "true")


# ====== INPUT FILE BOQ DOWNLOAD ======

INPUT_DIR_NAME = "input"
DEFAULT_INPUT_XLSX = "selected_tenders.xlsx"
MISC_BOQ_FOLDER = "Miscellaneous BoQs"
PORTAL_BASE_URL = "https://eproc.rajasthan.gov.in"


def get_input_dir():
    path = os.path.join(base_path, INPUT_DIR_NAME)
    os.makedirs(path, exist_ok=True)
    return path


def resolve_input_xlsx_path(filename):
    """Resolve filename under input/ (adds .xlsx if missing)."""
    name = (filename or DEFAULT_INPUT_XLSX).strip()
    if not name.lower().endswith(".xlsx"):
        name = f"{name}.xlsx"
    return os.path.join(get_input_dir(), name)


def _normalize_portal_tender_url(url):
    if not url:
        return None
    u = str(url).strip()
    if not u or u.lower() in ("nan", "none"):
        return None
    if u.startswith("http://") or u.startswith("https://"):
        return u
    if u.startswith("/"):
        return PORTAL_BASE_URL.rstrip("/") + u
    return u


def _cell_hyperlink_url(cell):
    hl = getattr(cell, "hyperlink", None)
    if hl is None:
        return None
    target = getattr(hl, "target", None) or getattr(hl, "location", None)
    if target:
        return _normalize_portal_tender_url(target)
    return None


def load_tenders_from_input_xlsx(xlsx_path):
    """
    Read tenders from an Excel file (same layout as exported *_tenders.xlsx).
    Uses Title hyperlinks and/or an explicit Tender Link column.
    Returns list of tender dicts with at least Title and Tender Link.
    """
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"Input file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None and str(val).strip():
            headers[str(val).strip()] = col

    if "Title" not in headers:
        wb.close()
        raise ValueError("Input Excel must have a 'Title' column (row 1).")

    optional_cols = [
        "Tender ID", "Organisation Chain", "Closing Date", "Amount",
        "Work Description", "Location", "Zipcode", "District", "State/U.T.",
    ]
    link_col = headers.get("Tender Link")

    tenders = []
    skipped = 0
    for row in range(2, ws.max_row + 1):
        title_cell = ws.cell(row=row, column=headers["Title"])
        title = title_cell.value
        title = "" if title is None else str(title).strip()
        if not title and link_col is None:
            continue

        link = None
        if link_col:
            raw_link = ws.cell(row=row, column=link_col).value
            link = _normalize_portal_tender_url(raw_link)
        if not link:
            link = _cell_hyperlink_url(title_cell)
        if not link:
            skipped += 1
            logging.warning("Row %s skipped (no tender link): %s", row, title[:80] or "(empty)")
            continue

        tender = {
            "Title": title or f"Tender row {row}",
            "Tender Link": link,
            "Organisation": "Miscellaneous (input file)",
        }
        for col_name in optional_cols:
            if col_name in headers:
                val = ws.cell(row=row, column=headers[col_name]).value
                if val is not None and str(val).strip() and str(val).lower() != "nan":
                    tender[col_name] = str(val).strip()
        tenders.append(tender)

    wb.close()
    if skipped:
        logging.info("Input file: skipped %s row(s) without a tender link.", skipped)
    return tenders


def prompt_run_mode():
    print(
        "\nWhat would you like to do?\n"
        "  1 — Fetch department-wise tenders (scrape portal)\n"
        "  2 — Download BOQs from input Excel file\n"
    )
    while True:
        raw = input("Choice [1/2]: ").strip().lower()
        if raw in ("1", "fetch", "dept", "tenders", "scrape"):
            return "fetch"
        if raw in ("2", "boq", "download", "input"):
            return "boq"
        print("Enter 1 or 2.\n")


def prompt_input_xlsx_filename():
    input_dir = get_input_dir()
    print(f"\nPlace your Excel file in:\n  {input_dir}\n")
    print(f"Expected columns match exported *_tenders.xlsx (Title with hyperlink is required).")
    print(f"Default file name: {DEFAULT_INPUT_XLSX}\n")
    raw = input(f"Input file name [{DEFAULT_INPUT_XLSX}]: ").strip()
    path = resolve_input_xlsx_path(raw or DEFAULT_INPUT_XLSX)
    if not os.path.isfile(path):
        print(f"\nFile not found:\n  {path}")
        print(f"Copy your .xlsx into the input folder and run again.\n")
        sys.exit(1)
    return path


def run_boq_from_input_file(config, callbacks=None):
    """
    Download BOQs for tenders listed in an input Excel file.
    Output: results/rajasthan/YYYY-MM-DD/BOQs/Miscellaneous BoQs/
    """
    global _captcha_callback
    cb = callbacks or _DefaultCallbacks()
    _captcha_callback = cb.on_captcha

    xlsx_path = config.get("input_xlsx_path") or resolve_input_xlsx_path(
        config.get("input_filename", DEFAULT_INPUT_XLSX)
    )
    cb.on_log(f"Reading input file: {xlsx_path}")
    tenders = load_tenders_from_input_xlsx(xlsx_path)
    if not tenders:
        cb.on_log("No tenders with valid links found in the input file.")
        summary = [{
            "org": "Miscellaneous (input file)",
            "total": 0,
            "filtered": 0,
            "boqs": 0,
        }]
        cb.on_complete(summary)
        return None, summary

    _assign_tender_ids(tenders)
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    output_root = os.path.join(base_path, "results", "rajasthan", date_str)
    boq_dir = os.path.join(output_root, "BOQs", MISC_BOQ_FOLDER)
    staging_dir = os.path.join(output_root, "BOQs", "_chrome_staging")
    os.makedirs(boq_dir, exist_ok=True)
    os.makedirs(staging_dir, exist_ok=True)

    cb.on_log(f"Output folder: {boq_dir}")
    cb.on_log(f"Downloading BOQs for {len(tenders)} tender(s)...")

    driver = init_driver(headless=False, download_dir=staging_dir)
    cb.on_log("Chrome opened for BOQ downloads.")
    try:
        driver = download_boq_for_tenders(
            tenders,
            boq_dir,
            driver,
            staging_dir,
            department_name="Miscellaneous",
        )
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    boq_count = sum(1 for t in tenders if t.get("BOQ_Status") == "ok")
    cb.on_log(f"BOQs downloaded: {boq_count}/{len(tenders)}")
    summary = [{
        "org": "Miscellaneous (input file)",
        "total": len(tenders),
        "filtered": len(tenders),
        "boqs": boq_count,
    }]
    cb.on_complete(summary)
    return output_root, summary


def _reference_pin_key_from_excel_cell(val):
    """Normalize PIN from reference sheet (often float) to a 6-digit string key."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        n = int(round(float(val)))
        return f"{n:06d}"
    except (ValueError, TypeError, OverflowError):
        m = re.search(r"\d{6}", str(val))
        return m.group(0) if m else None


def _scraped_zip_to_pin_key(zip_val):
    """Extract a 6-digit Indian PIN from scraped Zipcode text."""
    if zip_val is None or (isinstance(zip_val, float) and pd.isna(zip_val)):
        return None
    s = str(zip_val).strip()
    if not s or s.lower() == "not found":
        return None
    m = re.search(r"\b(\d{6})\b", s)
    if m:
        return m.group(1)
    digits = re.sub(r"\D", "", s)
    if len(digits) >= 6:
        return digits[:6]
    return None


def load_pin_code_lookup():
    """
    Load Sheet1 from List of Pin Codes of Rajasthan.xlsx.
    Returns dict: 6-digit PIN str -> (DISTRICT, STATE/U.T.).
    First row wins if the same PIN appears multiple times (multiple offices).
    """
    path = os.path.join(base_path, PIN_CODES_REFERENCE_XLSX)
    if not os.path.isfile(path):
        logging.warning(
            "PIN reference file not found (%s). District and State/U.T. will be blank.",
            PIN_CODES_REFERENCE_XLSX,
        )
        return {}
    try:
        ref = pd.read_excel(path, sheet_name=0)
    except Exception as e:
        logging.error("Could not read PIN reference file: %s", e)
        return {}
    need = ("PINCODE", "DISTRICT", "STATE/U.T.")
    missing = [c for c in need if c not in ref.columns]
    if missing:
        logging.error(
            "PIN reference sheet missing columns %s. Found: %s",
            missing,
            list(ref.columns),
        )
        return {}
    lookup = {}
    for _, row in ref.iterrows():
        key = _reference_pin_key_from_excel_cell(row["PINCODE"])
        if not key:
            continue
        if key in lookup:
            continue
        dist = row["DISTRICT"]
        st = row["STATE/U.T."]
        dist_s = "" if pd.isna(dist) else str(dist).strip()
        st_s = "" if pd.isna(st) else str(st).strip()
        lookup[key] = (dist_s, st_s)
    logging.info(
        "Loaded PIN lookup: %s unique PINs from %s",
        len(lookup),
        PIN_CODES_REFERENCE_XLSX,
    )
    return lookup


def enrich_df_with_district_state(df, pin_lookup):
    """Add District and State/U.T. from pin_lookup using Zipcode column."""
    df = df.copy()
    districts = []
    states = []
    for z in df["Zipcode"]:
        key = _scraped_zip_to_pin_key(z)
        if not key:
            districts.append("")
            states.append("")
            continue
        if key in pin_lookup:
            d, s = pin_lookup[key]
            districts.append(d)
            states.append(s)
        else:
            districts.append("Not Found")
            states.append("Not Found")
    df["District"] = districts
    df["State/U.T."] = states
    return df


def export_formatted_excel(df, path, sheet_name="Active Tenders"):
    """
    Write tenders to xlsx with header styling, zebra rows, borders, hyperlinked titles, column widths.
    df must include plain-text Title and Tender Link plus export columns.
    """
    cols_order = [
        "Tender ID",
        "Title",
        "Organisation Chain",
        "Closing Date",
        "Amount",
        "Work Description",
        "Location",
        "Zipcode",
        "District",
        "State/U.T.",
    ]
    for c in cols_order:
        if c not in df.columns:
            raise ValueError(f"Missing column: {c}")
    if "Tender Link" not in df.columns:
        raise ValueError("Missing column: Tender Link")

    titles = ["" if pd.isna(t) else str(t).strip() for t in df["Title"]]
    links = ["" if pd.isna(u) else str(u).strip() for u in df["Tender Link"]]
    write_df = df[cols_order].copy()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        write_df.to_excel(writer, sheet_name=sheet_name, index=True, index_label="Sno")

    wb = load_workbook(path)
    ws = wb[sheet_name]

    thin = Side(style="thin", color="B4B4B4")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    fill_alt = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
    fill_base = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    link_font = Font(color="0563C1", underline="single", size=11)
    body_font = Font(size=11)

    max_col = ws.max_column
    title_col_idx = 3  # Col A = Sno, B = Tender ID, C = Title

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = grid_border

    for r in range(2, ws.max_row + 1):
        row_fill = fill_alt if (r - 2) % 2 == 0 else fill_base
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = grid_border
            cell.fill = row_fill
            if c == title_col_idx:
                url = links[r - 2] if r - 2 < len(links) else ""
                tit = titles[r - 2] if r - 2 < len(titles) else ""
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.value = tit if tit and tit != "nan" else url
                    cell.font = link_font
                else:
                    cell.value = tit
                    cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 7,
        2: 12,
        3: 40,
        4: 32,
        5: 18,
        6: 14,
        7: 40,
        8: 18,
        9: 10,
        10: 18,
        11: 14,
    }
    for col_idx, w in widths.items():
        if col_idx <= max_col:
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True
    wb.save(path)


def build_export_dataframe(tenders):
    df = pd.DataFrame(tenders)
    if "Tender ID" not in df.columns:
        df["Tender ID"] = ""
    df = df.rename(
        columns={
            "Title": "Title",
            "Organisation Chain": "Organisation Chain",
            "Bid Submission End Date": "Closing Date",
            "Tender Value in ₹": "Amount",
            "Work Description": "Work Description",
            "Location": "Location",
        }
    )
    df["Tender ID"] = df["Tender ID"].fillna("").astype(str).str.strip()
    df["Title"] = df["Title"].fillna("").astype(str).str.strip()
    df["Title"] = df.apply(
        lambda r: f"{r['Tender ID']} - {r['Title']}" if r["Tender ID"] else r["Title"],
        axis=1,
    )
    df = df[
        [
            "Tender ID",
            "Title",
            "Tender Link",
            "Organisation Chain",
            "Closing Date",
            "Amount",
            "Work Description",
            "Location",
            "Zipcode",
        ]
    ]
    return df


# ====== PROGRAMMATIC API ======

class _DefaultCallbacks:
    """Fallback callbacks that print to stdout and use input() for CAPTCHA."""
    def on_log(self, message):
        print(message)

    def on_progress(self, current, total, label):
        pct = round(current / max(total, 1) * 100, 1)
        print(f"  [{current}/{total}] {pct}% — {label}")

    def on_captcha(self, label):
        return _prompt_boq_user_action_cli(label)

    def on_complete(self, summary):
        pass


def fetch_organisations():
    """Start a headless browser, navigate to eproc.rajasthan.gov.in, and return {org_name: anchor_id}."""
    driver = init_driver(headless=True)
    try:
        navigate_to_main_page(driver)
        org_dict = extract_orgs_and_anchor_ids_selenium(driver)
    finally:
        driver.quit()
    return org_dict


def run_scraper(config, callbacks=None):
    """
    Run the Rajasthan eProc scraper programmatically.

    config keys:
        org_jobs     — list of (org_name, anchor_id) tuples
        value_mode   — "all" or "range"
        range_min_r  — min rupees (when value_mode == "range")
        range_max_r  — max rupees (when value_mode == "range")
        download_boq — bool
    """
    global _captcha_callback
    cb = callbacks or _DefaultCallbacks()
    _captcha_callback = cb.on_captcha

    org_jobs = config["org_jobs"]
    value_mode = config.get("value_mode", "all")
    range_min_r = config.get("range_min_r")
    range_max_r = config.get("range_max_r")
    download_boq_flag = config.get("download_boq", False)

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    depwise_root = os.path.join(base_path, "results", "rajasthan", date_str)
    os.makedirs(depwise_root, exist_ok=True)

    cb.on_log(f"Output root: {depwise_root}")
    logging.info("Output root: %s", depwise_root)

    pin_lookup = load_pin_code_lookup()

    summary_results = []
    boq_staging_dir = os.path.join(depwise_root, "BOQs", "_chrome_staging")
    boq_driver = None

    try:
        for idx, (org_name, org_id) in enumerate(org_jobs):
            safe_dept = sanitize_folder_name(org_name)
            dept_dir = os.path.join(depwise_root, safe_dept)
            os.makedirs(dept_dir, exist_ok=True)
            excel_path = os.path.join(dept_dir, f"{safe_dept}_tenders.xlsx")

            cb.on_log(f"--- {org_name}")

            tenders = get_org_tenders(org_name, org_id)
            if not tenders:
                logging.warning("No tenders scraped for: %s", org_name)
                cb.on_log(f"  No tenders found for {org_name}")
                summary_results.append({"org": org_name, "total": 0, "filtered": 0, "boqs": 0})
                continue

            total_scraped = len(tenders)
            cb.on_progress(idx + 1, len(org_jobs), f"{org_name}: {total_scraped} tenders scraped")

            if value_mode == "range":
                before = len(tenders)
                tenders = filter_tenders_by_crore_range(tenders, range_min_r, range_max_r)
                cb.on_log(
                    f"  Value filter: {len(tenders)} of {before} in range; "
                    f"{before - len(tenders)} excluded."
                )
            if not tenders:
                summary_results.append({"org": org_name, "total": total_scraped, "filtered": 0, "boqs": 0})
                continue

            _assign_tender_ids(tenders)
            df = build_export_dataframe(tenders)
            df = enrich_df_with_district_state(df, pin_lookup)
            export_formatted_excel(df, excel_path)
            logging.info("Tenders exported to Excel: %s", excel_path)
            cb.on_log(f"  Excel saved: {excel_path}")

            boq_count = 0
            if download_boq_flag:
                if boq_driver is None:
                    os.makedirs(boq_staging_dir, exist_ok=True)
                    boq_driver = init_driver(
                        headless=False, download_dir=boq_staging_dir
                    )
                    cb.on_log("  Chrome opened for BOQ downloads.")
                boq_dept_dir = os.path.join(depwise_root, "BOQs", safe_dept)
                cb.on_log(f"  BOQ folder: {boq_dept_dir}")
                boq_driver = download_boq_for_tenders(
                    tenders, boq_dept_dir, boq_driver,
                    boq_staging_dir, department_name=org_name,
                )
                boq_count = sum(1 for t in tenders if t.get("BOQ_Status") == "ok")
                cb.on_log(f"  BOQs downloaded: {boq_count}/{len(tenders)}")

            summary_results.append({
                "org": org_name,
                "total": total_scraped,
                "filtered": len(tenders),
                "boqs": boq_count,
            })
    finally:
        if boq_driver is not None:
            try:
                boq_driver.quit()
            except Exception:
                pass

    cb.on_complete(summary_results)
    return depwise_root, summary_results


# ====== MAIN EXECUTION (CLI) ======
if __name__ == "__main__":
    run_mode = prompt_run_mode()

    if run_mode == "boq":
        input_path = prompt_input_xlsx_filename()
        _, summary = run_boq_from_input_file({"input_xlsx_path": input_path})
        if not summary or summary[0]["boqs"] == 0:
            total = summary[0]["total"] if summary else 0
            if total == 0:
                print("No tenders with links found in input file.")
            else:
                print("No BOQs downloaded successfully.")
            sys.exit(1)
        print("\nDone.")
        sys.exit(0)

    org_dict = fetch_organisations()

    if not org_dict:
        logging.error("No organisations found on the page.")
        sys.exit(1)

    print(f"\nFound {len(org_dict)} organisation(s) on the active tenders page.")
    print(
        "Type 'list' to print all names.\n"
        "Enter one organisation, or several separated by commas "
        "(exact name, partial match, or case-insensitive).\n"
    )

    while True:
        choice = input("Organisation(s) to scrape: ").strip()
        if choice.lower() == "list":
            for i, name in enumerate(sorted(org_dict.keys()), 1):
                print(f"  {i}. {name}")
            continue

        status, payload = resolve_org_list(org_dict, choice)
        if status == "error":
            print(payload)
            continue
        org_jobs = payload
        break

    value_mode, range_min_r, range_max_r = prompt_tender_value_filter()
    if value_mode == "range":
        lo_cr = range_min_r / RUPEES_PER_CRORE
        hi_cr = range_max_r / RUPEES_PER_CRORE
        print(f"\nFilter: ₹{lo_cr:g} Cr – ₹{hi_cr:g} Cr\n")
    else:
        print("\nNo value filter — including all amounts.\n")

    download_boq = prompt_download_boq()

    config = {
        "org_jobs": org_jobs,
        "value_mode": value_mode,
        "range_min_r": range_min_r,
        "range_max_r": range_max_r,
        "download_boq": download_boq,
    }

    _, summary = run_scraper(config)

    if not summary or all(s["filtered"] == 0 for s in summary):
        print("No tenders exported.")
        sys.exit(1)

    print("\nDone.")
