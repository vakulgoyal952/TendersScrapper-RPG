import io
import os
import re
import sys
import time
import shutil
import zipfile
import datetime
import logging
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
)
from webdriver_manager.chrome import ChromeDriverManager

PORTAL_URL = "https://etenders.gov.in/eprocure/app"
RUPEES_PER_CRORE = 10_000_000

TENDER_FIELDS = [
    "Title",
    "Organisation Chain",
    "Tender Value in \u20b9",
    "Work Description",
    "Location",
    "Bid Submission End Date",
]

TENDER_FIELD_ALIASES = {
    "Tender Value in \u20b9": [
        "Tender Value in \u20b9",
        "Tender Value in Rs.",
        "Tender Value",
        "Estimated Cost",
        "Tender Value in Rupees",
    ],
    "Bid Submission End Date": [
        "Bid Submission End Date",
        "Closing Date",
        "Bid Submission Closing Date",
        "Last Date for Submission",
    ],
}

ZIPCODE_LABELS = [
    "Pin Code", "Pincode", "PIN Code", "ZIP Code", "Zipcode", "Postal Code",
]


def get_base_path():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


base_path = get_base_path()


def setup_logging():
    log_dir = os.path.join(base_path, "logs", "national")
    os.makedirs(log_dir, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_file = os.path.join(log_dir, f"scraper_{ts}.log")
    logging.basicConfig(
        filename=log_file,
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    logging.getLogger().addHandler(console)
    logging.info("Logger ready \u2014 log file: %s", log_file)
    return log_file


# ====== DRIVER ======

def init_driver(headless=True, download_dir=None):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--start-maximized")
    options.page_load_strategy = "eager"
    if download_dir:
        os.makedirs(download_dir, exist_ok=True)
        dl = os.path.abspath(download_dir)
        options.add_experimental_option(
            "prefs",
            {
                "download.default_directory": dl,
                "download.prompt_for_download": False,
                "directory_upgrade": True,
                "safebrowsing.enabled": True,
            },
        )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


# ====== HELPERS ======

def sanitize_folder_name(name):
    name = (name or "").strip()
    for c in '<>:"/\\|?*\n\r\t':
        name = name.replace(c, "_")
    name = re.sub(r"_+", "_", name).strip("._ ")
    return name or "department"


def parse_inr_amount_to_rupees(text):
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return None
    s = str(text).strip()
    if not s or s.lower() in ("not found", "refer document", "refer tender document"):
        return None
    digits = re.sub(r"[^\d]", "", s)
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def sanitize_project_zip_basename(tender_data, seq_index):
    tender_id = str(tender_data.get("Tender ID") or "").strip()
    raw = (tender_data.get("Title") or "").strip()
    if not raw or raw.lower() == "not found":
        raw = (tender_data.get("Work Description") or "").strip()
    if not raw or raw.lower() == "not found":
        raw = f"tender_{seq_index:04d}"
    if tender_id:
        raw = f"{tender_id}_{raw}"
    base = sanitize_folder_name(raw)
    if len(base) > 150:
        base = base[:150].rstrip("_")
    return base or f"tender_{seq_index:04d}"


def unique_path_in_dir(target_dir, basename, ext=".zip"):
    p = os.path.join(target_dir, f"{basename}{ext}")
    if not os.path.exists(p):
        return p
    n = 2
    while True:
        p = os.path.join(target_dir, f"{basename}_{n}{ext}")
        if not os.path.exists(p):
            return p
        n += 1


def _assign_tender_ids(tenders):
    """Assign stable per-organisation IDs used in Excel and BOQ file names."""
    for i, t in enumerate(tenders, start=1):
        t["Tender ID"] = f"T{i:04d}"


def _tender_title_with_id(tender_data, fallback_seq):
    tid = str(tender_data.get("Tender ID") or "").strip() or f"T{fallback_seq:04d}"
    title = str(tender_data.get("Title") or "").strip()
    if title:
        return f"{tid}_{title}"
    return tid


# ====== NAVIGATION ======

def _wait_for_page_table(driver, timeout=15):
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(
                (By.XPATH,
                 "//table[.//tr[contains(@class,'even') or contains(@class,'odd')]]"
                 " | //table[@id='table']")
            )
        )
    except TimeoutException:
        pass


def navigate_to_org_listing(driver):
    driver.get(PORTAL_URL)
    try:
        link = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable(
                (By.XPATH,
                 "//a[contains(translate(normalize-space(.),"
                 " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                 " 'abcdefghijklmnopqrstuvwxyz'),"
                 " 'tender by organisation')]")
            )
        )
        link.click()
        _wait_for_page_table(driver)
        logging.info("Navigated to Tender by Organisation page.")
    except TimeoutException:
        driver.get(
            f"{PORTAL_URL}?page=FrontEndTendersByOrganisation&service=page"
        )
        _wait_for_page_table(driver)
        logging.info("Navigated to Tender by Organisation page via direct URL.")


def get_all_organisations(driver):
    org_dict = {}
    try:
        table = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "table"))
        )
    except TimeoutException:
        tables = driver.find_elements(By.TAG_NAME, "table")
        table = None
        for t in tables:
            if len(t.find_elements(By.TAG_NAME, "tr")) > 3:
                table = t
                break
        if table is None:
            logging.error("Could not find organisation table on the page.")
            return org_dict

    rows = table.find_elements(By.TAG_NAME, "tr")
    for row in rows:
        classes = row.get_attribute("class") or ""
        if "list_header" in classes:
            continue
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) >= 3:
            org_name = cells[1].text.strip()
            if not org_name:
                continue
            try:
                anchor = cells[2].find_element(By.TAG_NAME, "a")
                anchor_id = anchor.get_attribute("id")
                if anchor_id:
                    org_dict[org_name] = anchor_id
            except NoSuchElementException:
                pass
    logging.info("Found %d organisations on the listing page.", len(org_dict))
    return org_dict


def resolve_org_choice(org_dict, user_input):
    user_input = (user_input or "").strip()
    if not user_input:
        return None
    if user_input in org_dict:
        return (user_input, org_dict[user_input])
    lower_map = {k.lower(): k for k in org_dict}
    ul = user_input.lower()
    if ul in lower_map:
        k = lower_map[ul]
        return (k, org_dict[k])
    matches = [(k, org_dict[k]) for k in org_dict if ul in k.lower()]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        logging.warning("Ambiguous match for '%s': %s", user_input, [m[0] for m in matches])
        return matches[0]
    logging.warning("No organisation matched '%s'.", user_input)
    return None


def resolve_org_list(org_dict, raw_line):
    raw_line = (raw_line or "").strip()
    if not raw_line:
        return "error", "Empty input."
    tokens = [t.strip() for t in raw_line.split(",") if t.strip()]
    if not tokens:
        return "error", "No organisation names found (use commas between names)."
    resolved, seen = [], set()
    for token in tokens:
        match = resolve_org_choice(org_dict, token)
        if match is None:
            return "error", f"No organisation matched '{token}'. Type 'list' to see names."
        org_name, org_id = match
        if org_name not in seen:
            seen.add(org_name)
            resolved.append((org_name, org_id))
    if not resolved:
        return "error", "No valid organisations to scrape."
    return "ok", resolved


def click_organisation(driver, org_dict, org_name_query):
    match = resolve_org_choice(org_dict, org_name_query)
    if match is None:
        print(f"  No organisation matched '{org_name_query}'")
        return None
    canonical_name, anchor_id = match
    try:
        el = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, anchor_id))
        )
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", el
        )
        el.click()
        _wait_for_page_table(driver)
        logging.info("Clicked organisation: %s (id=%s)", canonical_name, anchor_id)
        return canonical_name
    except Exception as e:
        logging.error("Failed to click organisation %s: %s", canonical_name, e)
        return None


def collect_tender_links_current_page(driver):
    links = []
    try:
        table = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "table"))
        )
    except TimeoutException:
        tables = driver.find_elements(By.TAG_NAME, "table")
        table = None
        for t in tables:
            if len(t.find_elements(By.TAG_NAME, "tr")) > 2:
                table = t
                break
        if table is None:
            return links

    rows = table.find_elements(By.TAG_NAME, "tr")
    for row in rows:
        cls = row.get_attribute("class") or ""
        if "list_header" in cls:
            continue
        cells = row.find_elements(By.TAG_NAME, "td")
        for cell in cells:
            anchors = cell.find_elements(By.TAG_NAME, "a")
            for a in anchors:
                href = a.get_attribute("href") or ""
                aid = a.get_attribute("id") or ""
                if "DirectLink" in aid or "FrontEndViewTender" in href:
                    if href and href not in links:
                        links.append(href)
                    break
    return links


def collect_all_tender_links(driver):
    all_links = []
    page_num = 1
    while True:
        page_links = collect_tender_links_current_page(driver)
        all_links.extend(page_links)
        logging.info(
            "Page %d: collected %d tender links (total so far: %d)",
            page_num, len(page_links), len(all_links),
        )
        next_link = None
        for xpath in [
            "//a[contains(translate(normalize-space(.),"
            " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
            " 'abcdefghijklmnopqrstuvwxyz'), 'next')]",
            "//a[normalize-space(.)='>']",
        ]:
            try:
                for p in driver.find_elements(By.XPATH, xpath):
                    if p.is_displayed() and p.is_enabled():
                        next_link = p
                        break
            except Exception:
                pass
            if next_link:
                break
        if next_link is None:
            break
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", next_link
            )
            next_link.click()
            _wait_for_page_table(driver)
            page_num += 1
        except Exception as e:
            logging.warning("Could not click Next on page %d: %s", page_num, e)
            break
    logging.info("Total tender links collected: %d", len(all_links))
    return all_links


# ====== TENDER DETAIL SCRAPING ======

def _extract_all_fields_from_page(driver):
    result = {}
    try:
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.XPATH, "//td/b"))
        )
    except TimeoutException:
        return result
    for b in driver.find_elements(By.XPATH, "//td/b"):
        try:
            label = b.text.strip()
            if not label:
                continue
            parent_td = b.find_element(By.XPATH, "./ancestor::td")
            value_td = parent_td.find_element(
                By.XPATH, "./following-sibling::td[1]"
            )
            result[label] = value_td.text.strip()
        except Exception:
            continue
    return result


def _resolve_field(page_data, field_name):
    val = page_data.get(field_name)
    if val:
        return val
    for alias in TENDER_FIELD_ALIASES.get(field_name, []):
        val = page_data.get(alias)
        if val:
            return val
    return "Not Found"


def _recover_page_if_needed(driver, tender_url, max_attempts=3):
    for attempt in range(max_attempts):
        restart = None
        for xpath in [
            "//a[contains(translate(normalize-space(.),"
            " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
            " 'abcdefghijklmnopqrstuvwxyz'), 'restart')]",
            "//button[contains(translate(normalize-space(.),"
            " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
            " 'abcdefghijklmnopqrstuvwxyz'), 'restart')]",
        ]:
            for el in driver.find_elements(By.XPATH, xpath):
                try:
                    if el.is_displayed() and el.is_enabled():
                        restart = el
                        break
                except Exception:
                    continue
            if restart:
                break
        if restart is None:
            return
        logging.info(
            "Session/timeout \u2014 clicking Restart (attempt %d/%d)",
            attempt + 1, max_attempts,
        )
        try:
            restart.click()
            time.sleep(1)
            driver.get(tender_url)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//td/b"))
            )
        except Exception as e:
            logging.warning("Restart recovery failed: %s", e)


def scrape_tender_detail(driver, tender_url):
    try:
        driver.get(tender_url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//td/b"))
        )
    except TimeoutException:
        _recover_page_if_needed(driver, tender_url)
    except Exception as e:
        logging.error("Failed to load tender page %s: %s", tender_url, e)
        return None

    _recover_page_if_needed(driver, tender_url)
    page_data = _extract_all_fields_from_page(driver)
    tender_data = {"Tender Link": tender_url}
    for field in TENDER_FIELDS:
        tender_data[field] = _resolve_field(page_data, field)
    zipcode = "Not Found"
    for label in ZIPCODE_LABELS:
        val = page_data.get(label)
        if val:
            zipcode = val
            break
    tender_data["Zipcode"] = zipcode
    return tender_data


def _scrape_tender_worker(tender_url, org_name):
    driver = init_driver(headless=True)
    try:
        data = scrape_tender_detail(driver, tender_url)
        if data:
            data["Organisation"] = org_name
        return data
    except Exception as e:
        logging.error("Worker error for %s: %s", tender_url, e)
        return None
    finally:
        try:
            driver.quit()
        except Exception:
            pass


def scrape_all_tenders_parallel(tender_links, org_name, max_workers=4):
    tenders = []
    total = len(tender_links)
    completed = 0
    lock = threading.Lock()
    print(f"  Scraping {total} tenders with {max_workers} parallel workers...")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_url = {
            executor.submit(_scrape_tender_worker, url, org_name): url
            for url in tender_links
        }
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            with lock:
                completed += 1
            try:
                data = future.result()
                if data:
                    tenders.append(data)
                    pct = round(completed / total * 100, 1)
                    logging.info(
                        "[%d/%d] %.1f%% \u2014 %s",
                        completed, total, pct, (data.get("Title") or "")[:80],
                    )
                else:
                    logging.warning("[%d/%d] Failed: %s", completed, total, url)
            except Exception as e:
                logging.error(
                    "[%d/%d] Exception: %s \u2014 %s", completed, total, url, e
                )
            if completed % 20 == 0 or completed == total:
                print(
                    f"  [{completed}/{total}]"
                    f" {round(completed / total * 100, 1)}% done"
                )
    return tenders


def filter_tenders_by_value(tenders, min_r, max_r):
    kept, dropped_missing, dropped_range = [], 0, 0
    for t in tenders:
        raw = t.get("Tender Value in \u20b9", "")
        val = parse_inr_amount_to_rupees(raw)
        if val is None:
            dropped_missing += 1
            continue
        if val < min_r or val > max_r:
            dropped_range += 1
            continue
        kept.append(t)
    logging.info(
        "Value filter: kept %d, dropped %d (missing), %d (out of range)",
        len(kept), dropped_missing, dropped_range,
    )
    print(
        f"  Value filter: {len(kept)} kept,"
        f" {dropped_missing} missing value,"
        f" {dropped_range} out of range"
    )
    return kept


# ====== BOQ DOWNLOAD ======

def find_download_zip_link(driver):
    xpaths = [
        "//a[contains(translate(normalize-space(string(.)),"
        " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
        " 'abcdefghijklmnopqrstuvwxyz'), 'download a zip')]",
        "//a[contains(translate(normalize-space(string(.)),"
        " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
        " 'abcdefghijklmnopqrstuvwxyz'), 'zip file')]",
        "//a[contains(translate(normalize-space(string(.)),"
        " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
        " 'abcdefghijklmnopqrstuvwxyz'), 'download')"
        " and contains(translate(normalize-space(string(.)),"
        " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
        " 'abcdefghijklmnopqrstuvwxyz'), 'zip')]",
    ]
    for xp in xpaths:
        for a in driver.find_elements(By.XPATH, xp):
            try:
                if a.is_displayed() and a.is_enabled():
                    return a
            except Exception:
                continue
    for a in driver.find_elements(By.TAG_NAME, "a"):
        try:
            txt = (a.text or "").lower()
            if "zip" in txt and "download" in txt and a.is_displayed():
                return a
        except Exception:
            continue
    return None


def _safe_click(driver, element):
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", element
        )
        time.sleep(0.3)
        element.click()
        return True
    except Exception as e:
        logging.warning("Click failed: %s", e)
        return False


def wait_for_zip_download(download_dir, before_names, timeout=180):
    deadline = time.time() + timeout
    before_names = set(before_names)
    while time.time() < deadline:
        try:
            names = os.listdir(download_dir)
        except OSError:
            time.sleep(0.5)
            continue
        if any(n.endswith(".crdownload") for n in names):
            time.sleep(0.5)
            continue
        new_zips = [
            n for n in names if n.endswith(".zip") and n not in before_names
        ]
        if new_zips:
            new_zips.sort(
                key=lambda n: os.path.getmtime(
                    os.path.join(download_dir, n)
                ),
                reverse=True,
            )
            return os.path.join(download_dir, new_zips[0])
        time.sleep(0.5)
    return None


_captcha_callback = None

BOQ_ACTION_CONTINUE = "continue"
BOQ_ACTION_SKIP = "skip"
BOQ_ACTION_RESTART = "restart"
BOQ_RESTART_SENTINEL = "__BOQ_RESTART_BROWSER__"


def _normalize_boq_user_action(value):
    if value is True or value == BOQ_ACTION_CONTINUE:
        return BOQ_ACTION_CONTINUE
    if value == BOQ_ACTION_RESTART:
        return BOQ_ACTION_RESTART
    return BOQ_ACTION_SKIP


def _prompt_boq_user_action_cli(label, reason=""):
    reason_line = f"  Reason: {reason}\n" if reason else ""
    print(
        f"\n  >>> BOQ help needed — {label}\n"
        f"{reason_line}"
        "  [Enter] Continue — fix page/CAPTCHA, retry in same browser\n"
        "  [*]     Restart — close Chrome, fresh window, retry THIS tender\n"
        "  [s]     Skip this tender\n"
    )
    try:
        raw = input("  Choice [Enter/* / s]: ").strip().lower()
    except EOFError:
        return BOQ_ACTION_SKIP
    if raw in ("*", "r", "restart"):
        return BOQ_ACTION_RESTART
    if raw in ("s", "skip"):
        return BOQ_ACTION_SKIP
    return BOQ_ACTION_CONTINUE


def _ask_boq_user_action(label, reason=""):
    global _captcha_callback
    if _captcha_callback is not None:
        msg = f"{reason} — {label}" if reason else label
        return _normalize_boq_user_action(_captcha_callback(msg))
    return _prompt_boq_user_action_cli(label, reason)


def _restart_boq_driver(driver, staging_dir):
    try:
        driver.quit()
    except Exception:
        pass
    time.sleep(1)
    logging.info("Restarting Chrome for BOQ downloads (new browser session).")
    return init_driver(headless=False, download_dir=staging_dir)


def try_download_zip(driver, download_dir, tender_label, tender_url=None):
    link = find_download_zip_link(driver)
    if not link:
        logging.warning("No zip download link for: %s", tender_label)
        return None
    before = set(os.listdir(download_dir))
    if not _safe_click(driver, link):
        return None
    path = wait_for_zip_download(download_dir, before, timeout=25)
    if path:
        logging.info("ZIP downloaded (no CAPTCHA) for: %s", tender_label)
        return path

    for attempt in range(1, 4):
        action = _ask_boq_user_action(
            tender_label,
            reason=f"CAPTCHA or download retry ({attempt}/3)",
        )
        if action == BOQ_ACTION_RESTART:
            return BOQ_RESTART_SENTINEL
        if action == BOQ_ACTION_SKIP:
            logging.info("User skipped BOQ for: %s", tender_label)
            return None
        _recover_page_if_needed(driver, tender_url or "")
        link2 = find_download_zip_link(driver)
        if not link2:
            logging.warning(
                "No zip link after CAPTCHA wait (round %d): %s",
                attempt, tender_label,
            )
            if tender_url:
                driver.get(tender_url)
                time.sleep(2)
            continue
        before_round = set(os.listdir(download_dir))
        if not _safe_click(driver, link2):
            continue
        path = wait_for_zip_download(download_dir, before_round, timeout=180)
        if path:
            logging.info(
                "ZIP downloaded after CAPTCHA for: %s", tender_label
            )
            return path
        logging.warning(
            "No ZIP after click (round %d): %s", attempt, tender_label
        )
    logging.warning("Gave up on ZIP download for: %s", tender_label)
    return None


def extract_boq_from_zip(zip_path, boq_dir, tender_title, seq_index=0):
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            members = [n for n in zf.namelist() if not n.endswith("/")]
            if not members:
                return None, "empty_zip"
            base_names = [(n, os.path.basename(n)) for n in members]
            boq_files = [n for n, b in base_names if "boq" in b.lower()]
            candidates = boq_files or [
                n for n, b in base_names
                if b.lower().endswith((".xlsx", ".xls", ".csv"))
            ]
            if not candidates:
                return None, "no_boq_in_zip"
            inner = candidates[0]
            ext = os.path.splitext(inner)[1].lower()
            raw_bytes = zf.read(inner)
    except (zipfile.BadZipFile, Exception) as e:
        logging.error("Bad zip %s: %s", zip_path, e)
        return None, f"bad_zip: {e}"

    bio = io.BytesIO(raw_bytes)
    try:
        if ext == ".csv":
            df = pd.read_csv(bio, header=None, encoding_errors="replace")
        else:
            engine = "openpyxl" if ext == ".xlsx" else (
                "xlrd" if ext == ".xls" else None
            )
            kwargs = {"engine": engine} if engine else {}
            xls = pd.ExcelFile(bio, **kwargs)
            parts = [
                pd.read_excel(xls, sheet_name=s, header=None)
                for s in xls.sheet_names
            ]
            df = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
    except Exception as e:
        logging.error("Could not read BOQ from %s: %s", zip_path, e)
        return None, f"read_error: {e}"

    if df.empty:
        return None, "empty_boq"

    safe_title = sanitize_folder_name(tender_title or f"tender_{seq_index:04d}")
    if len(safe_title) > 120:
        safe_title = safe_title[:120].rstrip("_")
    pdf_path = unique_path_in_dir(boq_dir, safe_title, ext=".pdf")
    try:
        _write_boq_dataframe_to_pdf(df, pdf_path, tender_title or safe_title)
        logging.info("BOQ PDF: %s", pdf_path)
        return pdf_path, "ok"
    except Exception as e:
        logging.error("PDF write failed for %s: %s", tender_title, e)
        return None, f"pdf_error: {e}"


def _write_boq_dataframe_to_pdf(df, pdf_path, title):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )
    import html as html_module

    LM = RM = 14 * mm
    TM = BM = 12 * mm
    page_w, _ = landscape(A4)
    usable_w = page_w - LM - RM
    doc = SimpleDocTemplate(
        pdf_path, pagesize=landscape(A4),
        leftMargin=LM, rightMargin=RM, topMargin=TM, bottomMargin=BM,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "BoqTitle", parent=styles["Title"], fontSize=13, leading=16,
        textColor=colors.HexColor("#1F4E79"), spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "BoqCell", fontName="Helvetica", fontSize=7, leading=9,
        leftIndent=2, rightIndent=2, spaceBefore=1, spaceAfter=1,
        alignment=TA_LEFT,
    )
    header_style = ParagraphStyle(
        "BoqHead", parent=cell_style, fontName="Helvetica-Bold",
        fontSize=7, leading=9, textColor=colors.whitesmoke,
    )

    def safe_para(val, st):
        s = "" if pd.isna(val) else str(val).strip()
        s = html_module.escape(s).replace("\n", "<br/>")
        return Paragraph(s or " ", st)

    flow = [
        Paragraph(html_module.escape(title)[:200], title_style),
        Spacer(1, 4),
    ]
    ncols = df.shape[1]
    col_widths = [usable_w / max(ncols, 1)] * ncols

    header_labels = [f"Col {j}" for j in range(ncols)]
    for idx in range(min(5, len(df))):
        row_vals = [
            str(v).strip() if pd.notna(v) else "" for v in df.iloc[idx]
        ]
        if sum(1 for v in row_vals if v) >= ncols * 0.5:
            header_labels = row_vals
            df = df.iloc[idx + 1:].reset_index(drop=True)
            break

    table_data = [[safe_para(h, header_style) for h in header_labels]]
    for _, row in df.iterrows():
        table_data.append(
            [safe_para(row.iloc[j], cell_style) for j in range(ncols)]
        )
    t = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B0B0B0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F0F5FA")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    flow.append(t)
    doc.build(flow)


def download_boqs_for_tenders(tenders, boq_dir, staging_dir, driver):
    os.makedirs(boq_dir, exist_ok=True)
    total = len(tenders)
    for i, t in enumerate(tenders, start=1):
        t.setdefault("BOQ_Status", "")
        t.setdefault("BOQ_PDF", "")
        link = t.get("Tender Link")
        if not link:
            t["BOQ_Status"] = "no_link"
            continue
        label = f"{i}/{total} \u2014 {(t.get('Title') or '')[:60]}"
        tender_done = False
        while not tender_done:
            try:
                driver.get(link)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td/b"))
                )
            except Exception as e:
                logging.error("Failed to open tender %s: %s", label, e)
                action = _ask_boq_user_action(label, "Could not open tender page")
                if action == BOQ_ACTION_RESTART:
                    driver = _restart_boq_driver(driver, staging_dir)
                    continue
                if action == BOQ_ACTION_SKIP:
                    t["BOQ_Status"] = "page_error"
                    tender_done = True
                    continue
                continue

            _recover_page_if_needed(driver, link)
            zip_path = try_download_zip(
                driver, staging_dir, label, tender_url=link
            )
            if zip_path == BOQ_RESTART_SENTINEL:
                driver = _restart_boq_driver(driver, staging_dir)
                continue
            if not zip_path:
                action = _ask_boq_user_action(label, "ZIP download failed")
                if action == BOQ_ACTION_RESTART:
                    driver = _restart_boq_driver(driver, staging_dir)
                    continue
                if action == BOQ_ACTION_SKIP:
                    t["BOQ_Status"] = "download_failed"
                    tender_done = True
                    continue
                continue

            proj_base = sanitize_project_zip_basename(t, i)
            dest_zip = unique_path_in_dir(boq_dir, proj_base, ext=".zip")
            try:
                if os.path.abspath(zip_path) != os.path.abspath(dest_zip):
                    shutil.move(zip_path, dest_zip)
            except OSError as e:
                logging.warning("Could not move zip: %s", e)
                dest_zip = zip_path
            pdf_path, status = extract_boq_from_zip(
                dest_zip, boq_dir, _tender_title_with_id(t, i), seq_index=i,
            )
            t["BOQ_Status"] = status
            t["BOQ_PDF"] = pdf_path or ""
            if i % 5 == 0 or i == total:
                print(f"  BOQ [{i}/{total}] \u2014 status: {status}")
            tender_done = True
    return driver


# ====== EXCEL EXPORT ======

def build_export_dataframe(tenders):
    df = pd.DataFrame(tenders)
    if "Tender ID" not in df.columns:
        df["Tender ID"] = ""
    rename_map = {
        "Bid Submission End Date": "Closing Date",
        "Tender Value in \u20b9": "Amount",
    }
    df = df.rename(columns=rename_map)
    df["Tender ID"] = df["Tender ID"].fillna("").astype(str).str.strip()
    df["Title"] = df["Title"].fillna("").astype(str).str.strip()
    df["Title"] = df.apply(
        lambda r: f"{r['Tender ID']} - {r['Title']}" if r["Tender ID"] else r["Title"],
        axis=1,
    )
    required = [
        "Tender ID",
        "Title", "Tender Link", "Organisation Chain", "Closing Date",
        "Amount", "Work Description", "Location", "Zipcode",
    ]
    for col in required:
        if col not in df.columns:
            df[col] = ""
    if "District" not in df.columns:
        df["District"] = ""
    if "State/U.T." not in df.columns:
        df["State/U.T."] = ""
    for idx, row in df.iterrows():
        chain = str(row.get("Organisation Chain", ""))
        parts = [
            p.strip()
            for p in chain.replace("||", "|").split("|") if p.strip()
        ]
        if not row.get("State/U.T.") and len(parts) >= 3:
            df.at[idx, "State/U.T."] = (
                parts[-1] if len(parts[-1]) < 40 else ""
            )
        if not row.get("District") and len(parts) >= 4:
            df.at[idx, "District"] = (
                parts[-2] if len(parts[-2]) < 40 else ""
            )
    return df[required + ["District", "State/U.T."]]


def export_formatted_excel(df, path, sheet_name="Tenders"):
    cols_order = [
        "Tender ID", "Title", "Organisation Chain", "Closing Date", "Amount",
        "Work Description", "Location", "Zipcode", "District", "State/U.T.",
    ]
    for c in cols_order:
        if c not in df.columns:
            df[c] = ""
    titles = [
        ("" if pd.isna(t) else str(t).strip()) for t in df["Title"]
    ]
    links = [
        ("" if pd.isna(u) else str(u).strip()) for u in df["Tender Link"]
    ]
    write_df = df[cols_order].copy()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        write_df.to_excel(
            writer, sheet_name=sheet_name, index=True, index_label="Sno",
        )
    wb = load_workbook(path)
    ws = wb[sheet_name]
    thin = Side(style="thin", color="B4B4B4")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid",
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    fill_alt = PatternFill(
        start_color="E8F0F8", end_color="E8F0F8", fill_type="solid",
    )
    fill_base = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid",
    )
    link_font = Font(color="0563C1", underline="single", size=11)
    body_font = Font(size=11)
    max_col = ws.max_column
    title_col_idx = 3
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        cell.border = grid_border
    for r in range(2, ws.max_row + 1):
        row_fill = fill_alt if (r - 2) % 2 == 0 else fill_base
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = grid_border
            cell.fill = row_fill
            if c == title_col_idx:
                url = links[r - 2] if r - 2 < len(links) else ""
                tit = titles[r - 2] if r - 2 < len(titles) else ""
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.value = tit if tit and tit != "nan" else url
                    cell.font = link_font
                else:
                    cell.value = tit
                    cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    col_widths = {
        1: 7, 2: 12, 3: 40, 4: 32, 5: 18, 6: 16,
        7: 40, 8: 18, 9: 10, 10: 18, 11: 14,
    }
    for col_idx, w in col_widths.items():
        if col_idx <= max_col:
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    logging.info("Excel saved: %s", path)


# ====== INTERACTIVE PROMPTS ======

def prompt_organisations(org_dict):
    print(f"\nFound {len(org_dict)} organisation(s) on the portal.")
    print(
        "Type 'list' to print all names.\n"
        "Enter one organisation, or several separated by commas "
        "(exact name, partial match, or case-insensitive).\n"
    )
    while True:
        choice = input("Organisation(s) to scrape: ").strip()
        if choice.lower() == "list":
            for i, name in enumerate(sorted(org_dict.keys()), 1):
                print(f"  {i}. {name}")
            continue
        status, payload = resolve_org_list(org_dict, choice)
        if status == "error":
            print(payload)
            continue
        return payload


def prompt_tender_value_filter():
    print(
        "\nTender value filter:\n"
        "  Type  all  \u2014 keep every scraped tender (no value filter)\n"
        "  Or a range in crore:  8-50  means \u20b98 crore to \u20b950 crore\n"
    )
    while True:
        raw = input("Tender value (all or e.g. 8-50): ").strip()
        s = raw.lower().replace(" ", "")
        if s in ("all", "*", "any", "everything"):
            return "all", None, None
        m = re.match(r"^(\d+(?:\.\d+)?)-(\d+(?:\.\d+)?)$", s)
        if not m:
            print("Invalid. Use 'all' or two numbers like 8-50\n")
            continue
        lo, hi = float(m.group(1)), float(m.group(2))
        if lo <= 0 or hi <= 0:
            print("Values must be positive.\n")
            continue
        if lo > hi:
            lo, hi = hi, lo
        min_r = int(round(lo * RUPEES_PER_CRORE))
        max_r = int(round(hi * RUPEES_PER_CRORE))
        return "range", min_r, max_r


def prompt_download_boq():
    s = input("\nDownload BOQ ZIPs? [y/N]: ").strip().lower()
    return s in ("y", "yes", "1", "true")


def prompt_num_workers():
    raw = input("Parallel workers (default 4): ").strip()
    if not raw:
        return 4
    try:
        n = int(raw)
        return max(1, min(n, 8))
    except ValueError:
        return 4


# ====== PROGRAMMATIC API ======

class _DefaultCallbacks:
    """Fallback callbacks that print to stdout and use input() for CAPTCHA."""
    def on_log(self, message):
        print(message)

    def on_progress(self, current, total, label):
        pct = round(current / max(total, 1) * 100, 1)
        print(f"  [{current}/{total}] {pct}% \u2014 {label}")

    def on_captcha(self, label):
        return _prompt_boq_user_action_cli(label)

    def on_complete(self, summary):
        pass


def fetch_organisations():
    """Start a headless browser, navigate to etenders.gov.in, return {org_name: anchor_id}."""
    setup_logging()
    driver = init_driver(headless=True)
    try:
        navigate_to_org_listing(driver)
        org_dict = get_all_organisations(driver)
    finally:
        driver.quit()
    return org_dict


def run_scraper(config, callbacks=None):
    """
    Run the national eTenders scraper programmatically.

    config keys:
        org_jobs     \u2014 list of (org_name, anchor_id) tuples
        value_mode   \u2014 "all" or "range"
        range_min_r  \u2014 min rupees (when value_mode == "range")
        range_max_r  \u2014 max rupees (when value_mode == "range")
        download_boq \u2014 bool
        num_workers  \u2014 int (default 4)
    """
    global _captcha_callback
    cb = callbacks or _DefaultCallbacks()
    _captcha_callback = cb.on_captcha

    org_jobs = config["org_jobs"]
    value_mode = config.get("value_mode", "all")
    range_min_r = config.get("range_min_r")
    range_max_r = config.get("range_max_r")
    download_boq_flag = config.get("download_boq", False)
    num_workers = config.get("num_workers", 4)

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    output_root = os.path.join(base_path, "results", "national", date_str)
    os.makedirs(output_root, exist_ok=True)

    cb.on_log(f"Output root: {output_root}")
    logging.info("Output root: %s", output_root)

    # Phase 1 \u2014 Collect tender links
    cb.on_log("PHASE 1: Collecting tender links (headless)")
    driver = init_driver(headless=True)
    org_tender_links = {}
    try:
        for org_name, org_id in org_jobs:
            navigate_to_org_listing(driver)
            org_dict_fresh = get_all_organisations(driver)
            canonical = click_organisation(driver, org_dict_fresh, org_name)
            if canonical is None:
                cb.on_log(f"  Could not click '{org_name}' \u2014 skipping.")
                org_tender_links[org_name] = []
                continue
            links = collect_all_tender_links(driver)
            org_tender_links[org_name] = links
            cb.on_log(f"  {org_name}: {len(links)} tender link(s)")
    finally:
        driver.quit()

    # Phase 2 \u2014 Scrape details + export + BOQs
    cb.on_log(f"PHASE 2: Scraping details ({num_workers} workers, headless)")

    boq_driver = None
    boq_staging_dir = os.path.join(output_root, "_chrome_staging")
    summary_results = []

    try:
        for idx, (org_name, _) in enumerate(org_jobs):
            tender_links = org_tender_links.get(org_name, [])
            cb.on_log(f"--- {org_name} ({len(tender_links)} tenders)")

            if not tender_links:
                summary_results.append({"org": org_name, "total": 0, "filtered": 0, "boqs": 0})
                continue

            safe_name = sanitize_folder_name(org_name)
            org_dir = os.path.join(output_root, safe_name)
            boq_dir = os.path.join(org_dir, "BOQs")
            os.makedirs(boq_dir, exist_ok=True)
            excel_path = os.path.join(org_dir, f"{safe_name}_tenders.xlsx")

            tenders = scrape_all_tenders_parallel(
                tender_links, org_name, max_workers=num_workers,
            )
            total_scraped = len(tenders)
            cb.on_log(f"  Scraped {total_scraped} tender(s)")
            cb.on_progress(idx + 1, len(org_jobs), f"{org_name}: {total_scraped} tenders")

            if value_mode == "range":
                tenders = filter_tenders_by_value(tenders, range_min_r, range_max_r)
                if not tenders:
                    cb.on_log("  No tenders in range \u2014 skipping.")
                    summary_results.append({"org": org_name, "total": total_scraped, "filtered": 0, "boqs": 0})
                    continue

            if not tenders:
                summary_results.append({"org": org_name, "total": total_scraped, "filtered": 0, "boqs": 0})
                continue

            _assign_tender_ids(tenders)
            df = build_export_dataframe(tenders)
            export_formatted_excel(df, excel_path)
            cb.on_log(f"  Excel: {excel_path}")

            boq_count = 0
            if download_boq_flag:
                if boq_driver is None:
                    os.makedirs(boq_staging_dir, exist_ok=True)
                    boq_driver = init_driver(
                        headless=False, download_dir=boq_staging_dir,
                    )
                    cb.on_log("  Chrome opened for BOQ downloads.")
                cb.on_log(f"  BOQ folder: {boq_dir}")
                boq_driver = download_boqs_for_tenders(
                    tenders, boq_dir, boq_staging_dir, boq_driver,
                )
                boq_count = sum(
                    1 for t in tenders if t.get("BOQ_Status") == "ok"
                )
                cb.on_log(f"  BOQs: {boq_count}/{len(tenders)}")

            summary_results.append({
                "org": org_name,
                "total": total_scraped,
                "filtered": len(tenders),
                "boqs": boq_count,
            })
    finally:
        if boq_driver is not None:
            try:
                boq_driver.quit()
            except Exception:
                pass

    cb.on_complete(summary_results)
    return output_root, summary_results


# ====== MAIN EXECUTION (CLI) ======
if __name__ == "__main__":
    log_file = setup_logging()

    print("=" * 60)
    print("  National eTenders.gov.in Scraper")
    print("=" * 60)

    org_dict = fetch_organisations()

    if not org_dict:
        logging.error("No organisations found on the page.")
        sys.exit(1)

    org_jobs = prompt_organisations(org_dict)
    value_mode, range_min_r, range_max_r = prompt_tender_value_filter()
    download_boq = prompt_download_boq()
    num_workers = prompt_num_workers()

    if value_mode == "range":
        lo_cr = range_min_r / RUPEES_PER_CRORE
        hi_cr = range_max_r / RUPEES_PER_CRORE
        print(f"\nFilter: \u20b9{lo_cr:g} Cr \u2013 \u20b9{hi_cr:g} Cr")
    else:
        print("\nNo value filter \u2014 including all amounts.")

    config = {
        "org_jobs": org_jobs,
        "value_mode": value_mode,
        "range_min_r": range_min_r,
        "range_max_r": range_max_r,
        "download_boq": download_boq,
        "num_workers": num_workers,
    }

    _, summary = run_scraper(config)

    if not summary or all(s["filtered"] == 0 for s in summary):
        print("No tenders exported.")
        sys.exit(1)

    print("\nDone.")
